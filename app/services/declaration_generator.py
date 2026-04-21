"""
Генератор декларации по форме КНД 1152017 (УСН «Доходы»).

Формирует 4-страничный PDF, визуально близкий к официальной форме ФНС
(Приказ ФНС России от 02.10.2024 № ЕД-7-3/813@):

    Стр. 001 — Титульный лист
    Стр. 002 — Раздел 1.1
    Стр. 003 — Раздел 2.1.1
    Стр. 004 — Раздел 2.1.1 (продолжение, строки 140-143)

Также доступен XLSX-экспорт (рабочий вариант для редактирования).

Особенности:
- Разметка «в знакоместах»: каждая цифра/символ в отдельной клетке
- Сумы — в целых рублях (округление ROUND_HALF_UP — делает tax_engine)
- Поля, отсутствующие в проекте, оставляются пустыми (как в электронных формах)
- PDF417 в каждом заголовке заменён на заштрихованный прямоугольник-плейсхолдер
- Чёрные квадраты-маркеры в углах (как требует машинное распознавание ФНС)
"""

import io
import os
from decimal import Decimal
from datetime import datetime
from typing import Dict, Any, List, Optional

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# ШРИФТЫ
# =============================================================================
_FONT_REGISTERED = False
_FONT_REGULAR = "Helvetica"
_FONT_BOLD = "Helvetica-Bold"


def _register_cyrillic_font():
    """Регистрирует DejaVu (или другой шрифт с кириллицей)."""
    global _FONT_REGISTERED, _FONT_REGULAR, _FONT_BOLD
    if _FONT_REGISTERED:
        return _FONT_REGULAR, _FONT_BOLD

    paths_regular = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    paths_bold = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        "C:/Windows/Fonts/DejaVuSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]

    found_regular = None
    for p in paths_regular:
        if os.path.exists(p):
            found_regular = p
            break
    if not found_regular:
        _FONT_REGISTERED = True
        return _FONT_REGULAR, _FONT_BOLD  # fallback Helvetica

    pdfmetrics.registerFont(TTFont("DejaVu", found_regular))
    _FONT_REGULAR = "DejaVu"

    for p in paths_bold:
        if os.path.exists(p):
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", p))
            _FONT_BOLD = "DejaVu-Bold"
            break
    else:
        _FONT_BOLD = _FONT_REGULAR

    _FONT_REGISTERED = True
    return _FONT_REGULAR, _FONT_BOLD


# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ РЕНДЕРИНГА
# =============================================================================
if REPORTLAB_AVAILABLE:
    PAGE_W, PAGE_H = A4
    CELL_W = 5.0 * mm
    CELL_H = 6.5 * mm
else:
    PAGE_W, PAGE_H = 595.27, 841.89  # A4 in points (fallback)
    CELL_W = 14.17  # 5mm
    CELL_H = 18.43  # 6.5mm


def _int_str(value) -> str:
    """Целое число в виде строки, либо пусто если не задано."""
    if value is None or value == "":
        return ""
    try:
        return str(int(Decimal(str(value))))
    except Exception:
        return str(value)


def _draw_corner_markers(c: pdf_canvas.Canvas):
    """Чёрные квадраты в трёх углах страницы для ABBYY-распознавания."""
    size = 5 * mm
    offset = 6 * mm
    c.setFillColorRGB(0, 0, 0)
    # Левый верхний
    c.rect(offset, PAGE_H - offset - size, size, size, fill=1, stroke=0)
    # Правый верхний
    c.rect(PAGE_W - offset - size, PAGE_H - offset - size, size, size, fill=1, stroke=0)
    # Левый нижний
    c.rect(offset, offset, size, size, fill=1, stroke=0)
    c.setFillColorRGB(0, 0, 0)


def _draw_barcode_placeholder(c: pdf_canvas.Canvas, x: float, y: float, code: str):
    """
    Псевдо-штрихкод PDF417 (плейсхолдер).

    Реальный штрихкод ФНС печатается 1C/Контуром, но его воспроизведение
    в стороннем ПО некорректно. Вместо этого рисуем штриховой прямоугольник
    с человекочитаемым кодом формы (0301 5018/5025/5049/5056).
    """
    w = 32 * mm
    h = 12 * mm
    # Рамка
    c.setLineWidth(0.3)
    c.rect(x, y, w, h, stroke=1, fill=0)
    # Вертикальные штрихи (псевдо-код)
    c.setFillColorRGB(0, 0, 0)
    pattern = [1, 1, 0, 1, 0, 0, 1, 1, 1, 0, 1, 0, 0, 1, 0, 1,
               1, 0, 1, 1, 0, 1, 0, 1, 1, 1, 0, 0, 1, 0, 1, 1,
               0, 1, 1, 0, 1, 1, 0, 1, 0, 0, 1, 1, 0, 1, 1, 0]
    stripe_w = w / len(pattern)
    for i, p in enumerate(pattern):
        if p:
            c.rect(x + i * stripe_w, y + 1 * mm, stripe_w, h - 2 * mm, fill=1, stroke=0)
    # Код формы
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(x, y - 2.5 * mm, code)


def _draw_cells(
    c: pdf_canvas.Canvas,
    x: float,
    y: float,
    count: int,
    text: str = "",
    cell_w: float = CELL_W,
    cell_h: float = CELL_H,
    align: str = "left",
    font_size: float = 11,
    fill_empty_with_dash: bool = False,
):
    """
    Отрисовать ряд клеток (знакомест) с текстом.

    align="left" — стандарт ФНС: слева направо
    align="right" — если нужно (редко, для числовых правовыравненных полей)
    fill_empty_with_dash — поставить прочерк во всех пустых клетках
                           (актуально когда значение = 0 или пропущено)
    """
    c.setLineWidth(0.3)
    c.setStrokeColorRGB(0, 0, 0)
    for i in range(count):
        c.rect(x + i * cell_w, y, cell_w, cell_h, stroke=1, fill=0)

    text = (text or "").strip()
    if not text:
        if fill_empty_with_dash:
            c.setFont(_FONT_REGULAR, font_size)
            for i in range(count):
                tx = x + i * cell_w + cell_w / 2
                ty = y + cell_h / 2 - font_size * 0.35
                c.drawCentredString(tx, ty, "–")
        return

    chars = list(text)
    if len(chars) > count:
        chars = chars[:count]

    c.setFont(_FONT_REGULAR, font_size)
    if align == "right":
        start = count - len(chars)
    else:
        start = 0
    for i, ch in enumerate(chars):
        pos = start + i
        tx = x + pos * cell_w + cell_w / 2
        ty = y + cell_h / 2 - font_size * 0.35
        c.drawCentredString(tx, ty, ch)


def _draw_oktmo_cells(c: pdf_canvas.Canvas, x: float, y: float, oktmo: str):
    """
    ОКТМО — 11 знакомест. Если ОКТМО 8-значный, в оставшиеся пишутся прочерки.
    """
    o = str(oktmo or "").strip()
    # 11 cells
    c.setLineWidth(0.3)
    for i in range(11):
        c.rect(x + i * CELL_W, y, CELL_W, CELL_H, stroke=1, fill=0)
    chars = list(o)[:11]
    c.setFont(_FONT_REGULAR, 11)
    for i, ch in enumerate(chars):
        tx = x + i * CELL_W + CELL_W / 2
        ty = y + CELL_H / 2 - 4
        c.drawCentredString(tx, ty, ch)
    # Прочерки в пустых знакоместах (стандарт ФНС для 8-значного ОКТМО)
    for i in range(len(chars), 11):
        tx = x + i * CELL_W + CELL_W / 2
        ty = y + CELL_H / 2 - 4
        c.drawCentredString(tx, ty, "–")


def _draw_money_cells(c: pdf_canvas.Canvas, x: float, y: float, value):
    """
    Сумма в рублях в 10 знакоместах (нули и прочерки по соглашению ФНС).

    Если value is None или пусто — прочерки во всех 10 клетках.
    Иначе — сумма слева направо, остальные заполняются прочерками.
    """
    s = _int_str(value)
    c.setLineWidth(0.3)
    for i in range(10):
        c.rect(x + i * CELL_W, y, CELL_W, CELL_H, stroke=1, fill=0)
    c.setFont(_FONT_REGULAR, 11)
    if not s or s == "0":
        # Все 10 прочерков
        for i in range(10):
            tx = x + i * CELL_W + CELL_W / 2
            ty = y + CELL_H / 2 - 4
            c.drawCentredString(tx, ty, "–")
        return
    for i, ch in enumerate(s[:10]):
        tx = x + i * CELL_W + CELL_W / 2
        ty = y + CELL_H / 2 - 4
        c.drawCentredString(tx, ty, ch)
    for i in range(len(s), 10):
        tx = x + i * CELL_W + CELL_W / 2
        ty = y + CELL_H / 2 - 4
        c.drawCentredString(tx, ty, "–")


def _draw_rate_cells(c: pdf_canvas.Canvas, x: float, y: float, rate_x10):
    """
    Ставка налога в формате X . X (3 цифры до точки, 1 после = 4 знака + точка).

    Пример: rate_x10 = 60 → «6 . 0»
             rate_x10 = 80 → «8 . 0»
    """
    try:
        r = int(Decimal(str(rate_x10)))
    except Exception:
        r = 60
    integer_part = r // 10
    decimal_part = r % 10

    # 3 ячейки целая часть
    _draw_cells(c, x, y, 3, str(integer_part))
    # Точка (не ячейка — просто знак между)
    dot_x = x + 3 * CELL_W + 0.8 * mm
    c.setFont(_FONT_BOLD, 12)
    c.drawString(dot_x, y + CELL_H / 2 - 4, ".")
    # 2 ячейки после точки (обычно десятые: 0–9, сотые = 0)
    x2 = x + 3 * CELL_W + 2.5 * mm
    _draw_cells(c, x2, y, 2, f"{decimal_part}0")


def _draw_page_header(
    c: pdf_canvas.Canvas,
    inn: str,
    kpp: str,
    page_num: str,
    barcode_code: str,
):
    """
    Верхняя плашка страницы: штрихкод + ИНН/КПП + № страницы.
    """
    # Штрихкод (слева сверху, под угловой маркер)
    _draw_barcode_placeholder(c, 14 * mm, PAGE_H - 22 * mm, barcode_code)

    # Подпись ИНН / КПП / Стр. (мелкие буквы)
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(55 * mm, PAGE_H - 18 * mm, "ИНН")
    c.drawString(115 * mm, PAGE_H - 18 * mm, "КПП")
    c.drawString(170 * mm, PAGE_H - 18 * mm, "Стр.")

    # ИНН — 12 клеток
    _draw_cells(c, 55 * mm, PAGE_H - 26 * mm, 12, inn, cell_w=4.2 * mm, cell_h=5.5 * mm, font_size=10)
    # КПП — 9 клеток (для ИП всегда прочерки)
    kpp_chars = (kpp or "").strip()
    _draw_cells(
        c, 115 * mm, PAGE_H - 26 * mm, 9, kpp_chars,
        cell_w=4.2 * mm, cell_h=5.5 * mm, font_size=10,
        fill_empty_with_dash=not bool(kpp_chars),
    )
    # Стр. — 3 клетки
    _draw_cells(c, 170 * mm, PAGE_H - 26 * mm, 3, page_num.zfill(3),
                cell_w=4.2 * mm, cell_h=5.5 * mm, font_size=10)


def _draw_section_title(
    c: pdf_canvas.Canvas,
    y: float,
    title: str,
    subtitle: str = "",
):
    """Заголовок раздела, жирным по центру страницы."""
    c.setFont(_FONT_BOLD, 10)
    c.drawCentredString(PAGE_W / 2, y, title)
    if subtitle:
        c.setFont(_FONT_REGULAR, 8.5)
        # Разбиваем если длинный
        lines = _wrap_line(subtitle, 95)
        for i, line in enumerate(lines):
            c.drawCentredString(PAGE_W / 2, y - 4.5 * mm - i * 3.8 * mm, line)


def _wrap_line(text: str, max_chars: int) -> List[str]:
    """Простой перенос строк по словам."""
    words = text.split()
    lines = []
    cur = []
    cur_len = 0
    for w in words:
        if cur_len + len(w) + 1 > max_chars and cur:
            lines.append(" ".join(cur))
            cur = [w]
            cur_len = len(w)
        else:
            cur.append(w)
            cur_len += len(w) + 1
    if cur:
        lines.append(" ".join(cur))
    return lines


def _draw_field_label(c: pdf_canvas.Canvas, x: float, y: float, text: str, size: float = 7.5):
    """Метка поля — мелкий шрифт."""
    c.setFont(_FONT_REGULAR, size)
    c.drawString(x, y, text)


def _draw_signature_block(c: pdf_canvas.Canvas, y_top: float):
    """
    Блок «Достоверность и полноту сведений подтверждаю» внизу каждой страницы
    разделов 1.1 / 2.1.1.
    """
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, y_top, "Достоверность и полноту сведений, указанных на данной странице, подтверждаю:")
    # Подпись ________
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, y_top - 8 * mm, "Подпись ______________")
    c.drawString(85 * mm, y_top - 8 * mm, "Дата")
    # Дата: _ _ . _ _ . _ _ _ _
    _draw_cells(c, 100 * mm, y_top - 11 * mm, 2, "", cell_w=4.5 * mm, cell_h=5.5 * mm)
    c.setFont(_FONT_BOLD, 10)
    c.drawString(100 * mm + 9 * mm + 0.5 * mm, y_top - 10 * mm, ".")
    _draw_cells(c, 100 * mm + 10.5 * mm, y_top - 11 * mm, 2, "", cell_w=4.5 * mm, cell_h=5.5 * mm)
    c.drawString(100 * mm + 20 * mm, y_top - 10 * mm, ".")
    _draw_cells(c, 100 * mm + 21.5 * mm, y_top - 11 * mm, 4, "", cell_w=4.5 * mm, cell_h=5.5 * mm)


# =============================================================================
# СТРАНИЦА 001 — ТИТУЛЬНЫЙ ЛИСТ
# =============================================================================
def _draw_title_page(
    c: pdf_canvas.Canvas,
    inn: str,
    kpp: str,
    fio: str,
    year: str,
    ifns: str,
    phone: str,
    correction: str = "0--",
    period_code: str = "34",
    location_code: str = "120",
    pages_count: int = 4,
):
    """Страница 001 — Титульный лист декларации КНД 1152017."""
    _draw_corner_markers(c)
    _draw_page_header(c, inn, kpp, "001", "0301 5018")

    # Форма по КНД 1152017 (сверху справа)
    c.setFont(_FONT_REGULAR, 7)
    c.drawRightString(PAGE_W - 13 * mm, PAGE_H - 30 * mm, "Форма по КНД 1152017")

    # Заголовок: "Налоговая декларация по налогу, уплачиваемому..."
    c.setFont(_FONT_BOLD, 10)
    c.drawCentredString(PAGE_W / 2, PAGE_H - 40 * mm, "Налоговая декларация по налогу, уплачиваемому")
    c.setFont(_FONT_REGULAR, 10)
    c.drawCentredString(
        PAGE_W / 2, PAGE_H - 45 * mm,
        "в связи с применением упрощенной системы налогообложения"
    )

    # Номер корректировки (код) + Налоговый период (код) + Отчетный год
    y = PAGE_H - 55 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(15 * mm, y + 6 * mm, "Номер корректировки")
    _draw_cells(c, 15 * mm, y, 3, "0--", cell_w=4.8 * mm)

    # Разнесены шире, чтобы подписи не налезали друг на друга
    c.drawString(65 * mm, y + 6 * mm, "Налоговый период (код)")
    _draw_cells(c, 65 * mm, y, 2, period_code, cell_w=4.8 * mm)

    c.drawString(120 * mm, y + 6 * mm, "Отчетный год")
    _draw_cells(c, 120 * mm, y, 4, str(year), cell_w=4.8 * mm)

    # Представляется в налоговый орган (код) + по месту нахождения (код)
    y2 = y - 13 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(15 * mm, y2 + 6 * mm, "Представляется в налоговый орган (код)")
    _draw_cells(c, 15 * mm, y2, 4, ifns, cell_w=4.8 * mm,
                fill_empty_with_dash=not bool((ifns or "").strip()))

    c.drawString(90 * mm, y2 + 6 * mm, "по месту нахождения (учета) (код)")
    _draw_cells(c, 155 * mm, y2, 3, location_code, cell_w=4.8 * mm)

    # ФИО налогоплательщика (3 строки по 20 ячеек, UPPER CASE)
    # Подпись «(налогоплательщик)» размещается ВЫШЕ сетки, чтобы не налезать на буквы.
    y3 = y2 - 10 * mm
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(15 * mm, y3 + 6 * mm, "(налогоплательщик)")

    fio_upper = (fio or "").upper()
    fio_parts = fio_upper.split()
    fam = fio_parts[0] if len(fio_parts) > 0 else ""
    name = fio_parts[1] if len(fio_parts) > 1 else ""
    otch = " ".join(fio_parts[2:]) if len(fio_parts) > 2 else ""

    cells_per_fio = 20
    cw_fio = 4.5 * mm
    _draw_cells(c, 15 * mm, y3, cells_per_fio, fam, cell_w=cw_fio, cell_h=5.5 * mm, font_size=9)
    _draw_cells(c, 15 * mm, y3 - 6 * mm, cells_per_fio, name, cell_w=cw_fio, cell_h=5.5 * mm, font_size=9)
    _draw_cells(c, 15 * mm, y3 - 12 * mm, cells_per_fio, otch, cell_w=cw_fio, cell_h=5.5 * mm, font_size=9)

    # Форма реорганизации (ликвидация) + ИНН/КПП реорганизованной
    # Разносим подписи и сетки по разным координатам, чтобы не было наложения
    y4 = y3 - 22 * mm
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(15 * mm, y4 + 6 * mm, "Форма реорганизации (ликвидация) (код)")
    _draw_cells(c, 60 * mm, y4, 1, "", cell_w=4.8 * mm, fill_empty_with_dash=True)

    c.drawString(80 * mm, y4 + 6 * mm, "ИНН/КПП реорганизованной организации")
    # ИНН 10 клеток × 3.5mm = 35mm, / , КПП 9 клеток × 3.5mm = 31.5mm. Итого ~70mm.
    inn_cw = 3.5 * mm
    inn_x = 80 * mm
    _draw_cells(c, inn_x, y4, 10, "", cell_w=inn_cw, fill_empty_with_dash=True)
    slash_x = inn_x + 10 * inn_cw + 0.5 * mm
    c.setFont(_FONT_BOLD, 10)
    c.drawString(slash_x, y4 + CELL_H / 2 - 3, "/")
    c.setFont(_FONT_REGULAR, 7)
    _draw_cells(c, slash_x + 2.5 * mm, y4, 9, "", cell_w=inn_cw, fill_empty_with_dash=True)

    # Номер контактного телефона
    y5 = y4 - 13 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(15 * mm, y5 + 6 * mm, "Номер контактного телефона")
    cells_phone = 20
    _draw_cells(c, 15 * mm, y5, cells_phone, phone or "", cell_w=4.5 * mm, cell_h=5.5 * mm, font_size=9,
                fill_empty_with_dash=not bool((phone or "").strip()))

    # Объект налогообложения
    y6 = y5 - 13 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(15 * mm, y6 + 6 * mm, "Объект налогообложения:")
    _draw_cells(c, 100 * mm, y6, 1, "1", cell_w=4.8 * mm)
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(108 * mm, y6 + 3 * mm, "1 – доходы   /   2 – доходы, уменьшенные на величину расходов")

    # На ... страницах с приложением...
    y7 = y6 - 13 * mm
    c.setFont(_FONT_REGULAR, 6.5)
    c.drawString(15 * mm, y7 + 3 * mm, "На")
    _draw_cells(c, 22 * mm, y7, 3, str(pages_count).zfill(3), cell_w=4.2 * mm)
    c.drawString(37 * mm, y7 + 3 * mm,
                 "страницах с приложением подтверждающих документов")
    c.drawString(37 * mm, y7 - 1 * mm,
                 "или их копий на")
    _draw_cells(c, 67 * mm, y7 - 3 * mm, 3, "", cell_w=4.2 * mm, fill_empty_with_dash=True)
    c.drawString(82 * mm, y7 - 1 * mm, "листах")

    # ========== НИЖНЯЯ ЧАСТЬ — 2 СТОЛБЦА ==========
    y_bottom = y7 - 25 * mm

    # ЛЕВЫЙ СТОЛБЕЦ (уже — до 105 мм, чтобы не налезать на правый)
    col_left_x = 15 * mm
    col_left_w = 90 * mm
    c.setFont(_FONT_BOLD, 7.5)
    c.drawString(col_left_x, y_bottom + 10 * mm, "Достоверность и полноту сведений, указанных в")
    c.drawString(col_left_x, y_bottom + 6 * mm, "настоящей декларации, подтверждаю:")

    # Признак подписанта
    _draw_cells(c, col_left_x, y_bottom - 2 * mm, 1, "1", cell_w=4.8 * mm)
    c.setFont(_FONT_REGULAR, 7)
    c.drawString(col_left_x + 8 * mm, y_bottom - 1 * mm,
                 "1 - налогоплательщик / 2 - представитель налогоплательщика")

    # ФИО подписанта (3 строки)
    y_sign = y_bottom - 10 * mm
    _draw_cells(c, col_left_x, y_sign, 20, "", cell_w=3.8 * mm, cell_h=5 * mm, fill_empty_with_dash=True)
    _draw_cells(c, col_left_x, y_sign - 6 * mm, 20, "", cell_w=3.8 * mm, cell_h=5 * mm, fill_empty_with_dash=True)
    _draw_cells(c, col_left_x, y_sign - 12 * mm, 20, "", cell_w=3.8 * mm, cell_h=5 * mm, fill_empty_with_dash=True)

    c.setFont(_FONT_REGULAR, 6.5)
    c.drawString(col_left_x, y_sign - 16 * mm, "(фамилия, имя, отчество* полностью)")

    # Подпись и дата
    y_sign_date = y_sign - 23 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(col_left_x, y_sign_date, "Подпись ____________________")
    c.drawString(col_left_x + 40 * mm, y_sign_date, "Дата [__.__.__]")

    # Наименование и реквизиты документа
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(col_left_x, y_sign_date - 8 * mm,
                 "Наименование и реквизиты документа, подтверждающего полномочия")
    c.drawString(col_left_x, y_sign_date - 11 * mm, "представителя налогоплательщика")
    c.line(col_left_x, y_sign_date - 13 * mm, col_left_x + 95 * mm, y_sign_date - 13 * mm)

    c.setFont(_FONT_REGULAR, 6)
    c.drawString(col_left_x, y_bottom - 115 * mm, "*Отчество указывается при наличии.")

    # ПРАВЫЙ СТОЛБЕЦ
    col_right_x = 115 * mm
    col_right_w = 80 * mm
    c.setFont(_FONT_BOLD, 7.5)
    c.drawString(col_right_x, y_bottom + 10 * mm, "Заполняется работником налогового органа")

    c.setFont(_FONT_BOLD, 7)
    c.drawString(col_right_x, y_bottom + 5 * mm, "Сведения о представлении декларации")

    c.setFont(_FONT_REGULAR, 7)
    # Клетки для кода под лейблом, а не справа, чтобы не было наезда на текст
    c.drawString(col_right_x, y_bottom, "Данная декларация представлена (код)")
    _draw_cells(c, col_right_x + 62 * mm, y_bottom - 1 * mm, 3,
                "", cell_w=3.8 * mm, cell_h=4.5 * mm, fill_empty_with_dash=True)

    c.drawString(col_right_x, y_bottom - 8 * mm, "на")
    _draw_cells(c, col_right_x + 5 * mm, y_bottom - 10 * mm, 3,
                "", cell_w=3.8 * mm, cell_h=4.5 * mm, fill_empty_with_dash=True)
    c.drawString(col_right_x + 19 * mm, y_bottom - 8 * mm, "страницах")

    c.drawString(col_right_x, y_bottom - 16 * mm, "с приложением подтверждающих")
    c.drawString(col_right_x, y_bottom - 19 * mm, "документов или их копий на")
    _draw_cells(c, col_right_x + 43 * mm, y_bottom - 21 * mm, 3,
                "", cell_w=3.8 * mm, cell_h=4.5 * mm, fill_empty_with_dash=True)
    c.drawString(col_right_x + 57 * mm, y_bottom - 19 * mm, "листах")

    c.drawString(col_right_x, y_bottom - 28 * mm, "Дата представления декларации")
    _draw_cells(c, col_right_x, y_bottom - 34 * mm, 2,
                "", cell_w=3.8 * mm, cell_h=5 * mm)
    c.setFont(_FONT_BOLD, 10)
    c.drawString(col_right_x + 8 * mm, y_bottom - 33 * mm, ".")
    c.setFont(_FONT_REGULAR, 7)
    _draw_cells(c, col_right_x + 11 * mm, y_bottom - 34 * mm, 2,
                "", cell_w=3.8 * mm, cell_h=5 * mm)
    c.setFont(_FONT_BOLD, 10)
    c.drawString(col_right_x + 19 * mm, y_bottom - 33 * mm, ".")
    c.setFont(_FONT_REGULAR, 7)
    _draw_cells(c, col_right_x + 22 * mm, y_bottom - 34 * mm, 4,
                "", cell_w=3.8 * mm, cell_h=5 * mm)

    # Фамилия, И.О. подписанта ФНС
    y_fns_sign = y_bottom - 50 * mm
    c.setFont(_FONT_REGULAR, 7.5)
    c.drawString(col_right_x, y_fns_sign, "Фамилия, И.О.*")
    c.line(col_right_x + 25 * mm, y_fns_sign, col_right_x + col_right_w, y_fns_sign)

    c.drawString(col_right_x, y_fns_sign - 8 * mm, "Подпись")
    c.line(col_right_x + 25 * mm, y_fns_sign - 8 * mm, col_right_x + col_right_w, y_fns_sign - 8 * mm)


# =============================================================================
# СТРАНИЦА 002 — РАЗДЕЛ 1.1
# =============================================================================

# =============================================================================
# СТРАНИЦА 002 — РАЗДЕЛ 1.1
# =============================================================================
def _draw_section_1_1(
    c: pdf_canvas.Canvas,
    inn: str,
    kpp: str,
    s11: Dict[str, Any],
):
    """Раздел 1.1. Сумма налога (авансового платежа по налогу)."""
    _draw_corner_markers(c)
    _draw_page_header(c, inn, kpp, "002", "0301 5025")

    c.setFont(_FONT_BOLD, 9)
    title = "Раздел 1.1. Сумма налога (авансового платежа по налогу), уплачиваемого в связи с"
    c.drawCentredString(PAGE_W / 2, PAGE_H - 38 * mm, title)
    c.drawCentredString(PAGE_W / 2, PAGE_H - 42 * mm, "применением упрощенной системы налогообложения (объект налогообложения - доходы),")
    c.drawCentredString(PAGE_W / 2, PAGE_H - 46 * mm, "подлежащая уплате (уменьшению), по данным налогоплательщика")

    y_header = PAGE_H - 54 * mm
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, y_header, "Показатели")
    c.drawString(95 * mm, y_header, "Код строки")
    c.drawString(130 * mm, y_header, "Значения показателей (в рублях)")

    rows = [("Код по ОКТМО", "010", s11.get("line_010", ""), "oktmo"), ("Сумма авансового платежа, подлежащая уплате в срок не позднее\nдвадцать восьмого апреля отчетного года", "020", s11.get("line_020"), "money"), ("Код по ОКТМО", "030", s11.get("line_030", ""), "oktmo"), ("Сумма авансового платежа, подлежащая уплате в срок не позднее\nдвадцать восьмого июля отчетного года", "040", s11.get("line_040"), "money"), ("Сумма авансового платежа к уменьшению по сроку не позднее\nдвадцать восьмого июля отчетного года", "050", s11.get("line_050"), "money"), ("Код по ОКТМО", "060", s11.get("line_060", ""), "oktmo"), ("Сумма авансового платежа, подлежащая уплате в срок не позднее\nдвадцать восьмого октября отчетного года", "070", s11.get("line_070"), "money"), ("Сумма авансового платежа к уменьшению по сроку не позднее\nдвадцать восьмого октября отчетного года", "080", s11.get("line_080"), "money"), ("Код по ОКТМО", "090", s11.get("line_090", ""), "oktmo"), ("Сумма налога, подлежащая доплате за налоговый период\n(календарный год)", "100", s11.get("line_100"), "money"), ("Сумма налога, уплаченная в связи с применением ПСН,\nподлежащая зачету", "101", s11.get("line_101", ""), "money"), ("Сумма налога к уменьшению за налоговый период\n(календарный год)", "110", s11.get("line_110"), "money")]

    # Перенос длинных подписей по словам, чтобы они не налезали на колонку «Код строки».
    # Колонка подписей: 15 мм – 90 мм (~75 мм ≈ 52 символа при 7pt).
    y = y_header - 10 * mm
    for label, code, value, typ in rows:
        c.setFont(_FONT_REGULAR, 7)
        raw_lines: List[str] = []
        for part in label.split("\n"):
            raw_lines.extend(_wrap_line(part, 52))
        for j, line in enumerate(raw_lines):
            c.drawString(15 * mm, y - j * 3.5 * mm, line)
        _draw_cells(c, 95 * mm, y - 2 * mm, 3, code, cell_w=4.5 * mm, cell_h=5.5 * mm, font_size=10)
        if typ == "oktmo":
            _draw_oktmo_cells(c, 125 * mm, y - 2 * mm, str(value or ""))
        else:
            _draw_money_cells(c, 125 * mm, y - 2 * mm, value)
        y -= max(len(raw_lines) * 3.5 + 3, 8) * mm

    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, 30 * mm, "Достоверность и полноту сведений, указанных на данной странице, подтверждаю:")
    c.drawString(15 * mm, 20 * mm, "Подпись _______________________")
    c.drawString(90 * mm, 20 * mm, "Дата [__.__.__]")


# =============================================================================
# СТРАНИЦА 003 — РАЗДЕЛ 2.1.1 (часть 1)
# =============================================================================
def _draw_section_2_1_1_part1(
    c: pdf_canvas.Canvas,
    inn: str,
    kpp: str,
    s211: Dict[str, Any],
):
    """Раздел 2.1.1. Часть 1 — строки 101-133."""
    _draw_corner_markers(c)
    _draw_page_header(c, inn, kpp, "003", "0301 5049")

    c.setFont(_FONT_BOLD, 9)
    c.drawCentredString(PAGE_W / 2, PAGE_H - 38 * mm, "Раздел 2.1.1. Расчет налога, уплачиваемого в связи с применением")
    c.drawCentredString(PAGE_W / 2, PAGE_H - 42 * mm, "упрощенной системы налогообложения, (объект налогообложения – доходы)")

    y_header = PAGE_H - 50 * mm
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, y_header, "Показатели")
    c.drawString(95 * mm, y_header, "Код строки")
    c.drawString(130 * mm, y_header, "Значения показателей")

    rows = [("Код признака применения налоговой ставки", "101", s211.get("line_101", 1), "single"), ("Признак налогоплательщика", "102", s211.get("line_102", 2), "single"), ("Сумма полученных доходов за первый квартал (нарастающим итогом)", "110", s211.get("line_110"), "money9"), ("Сумма полученных доходов за полугодие (нарастающим итогом)", "111", s211.get("line_111"), "money9"), ("Сумма полученных доходов за девять месяцев (нарастающим итогом)", "112", s211.get("line_112"), "money9"), ("Сумма полученных доходов за налоговый период (нарастающим итогом)", "113", s211.get("line_113"), "money9"), ("Налоговая ставка (%) за первый квартал", "120", s211.get("line_120"), "rate"), ("Налоговая ставка (%) за полугодие", "121", s211.get("line_121"), "rate"), ("Налоговая ставка (%) за девять месяцев", "122", s211.get("line_122"), "rate"), ("Налоговая ставка (%) за налоговый период", "123", s211.get("line_123"), "rate"), ("Обоснование применения налоговой ставки, установленной законом\nсубъекта РФ", "124", s211.get("line_124", ""), "code"), ("Сумма исчисленного налога за первый квартал", "130", s211.get("line_130"), "money"), ("Сумма исчисленного налога за полугодие", "131", s211.get("line_131"), "money"), ("Сумма исчисленного налога за девять месяцев", "132", s211.get("line_132"), "money"), ("Сумма исчисленного налога за налоговый период", "133", s211.get("line_133"), "money")]

    y = y_header - 8 * mm
    for label, code, value, typ in rows:
        c.setFont(_FONT_REGULAR, 7)
        raw_lines: List[str] = []
        for part in label.split("\n"):
            raw_lines.extend(_wrap_line(part, 52))
        for j, line in enumerate(raw_lines):
            c.drawString(15 * mm, y - j * 3.5 * mm, line)
        _draw_cells(c, 95 * mm, y - 1 * mm, 3, code, cell_w=4.5 * mm, cell_h=5.5 * mm, font_size=9)
        if typ == "single":
            _draw_cells(c, 130 * mm, y - 1 * mm, 1, str(value), cell_w=5 * mm, font_size=10)
        elif typ == "rate":
            _draw_rate_cells(c, 130 * mm, y - 1 * mm, value)
        elif typ == "code":
            _draw_cells(c, 130 * mm, y - 1 * mm, 7, str(value or ""), cell_w=4.2 * mm)
            c.drawString(130 * mm + 7.5 * 4.2 * mm + 0.5 * mm, y - 1 * mm + CELL_H / 2 - 3, "/")
            _draw_cells(c, 130 * mm + 8.5 * 4.2 * mm + 1 * mm, y - 1 * mm, 12, "", cell_w=4.2 * mm, fill_empty_with_dash=True)
        elif typ == "money9":
            _draw_cells(c, 130 * mm, y - 1 * mm, 9, str(value or ""), cell_w=4.3 * mm, font_size=9)
        else:
            _draw_money_cells(c, 130 * mm, y - 1 * mm, value)
        y -= max(len(raw_lines) * 3.5 + 3, 8) * mm

    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, 30 * mm, "Достоверность и полноту сведений, указанных на данной странице, подтверждаю:")
    c.drawString(15 * mm, 20 * mm, "Подпись _______________________")
    c.drawString(90 * mm, 20 * mm, "Дата [__.__.__]")


# =============================================================================
# СТРАНИЦА 004 — РАЗДЕЛ 2.1.1 (часть 2)
# =============================================================================
def _draw_section_2_1_1_part2(
    c: pdf_canvas.Canvas,
    inn: str,
    kpp: str,
    s211: Dict[str, Any],
):
    """Раздел 2.1.1. Часть 2 — строки 140-143."""
    _draw_corner_markers(c)
    _draw_page_header(c, inn, kpp, "004", "0301 5056")

    c.setFont(_FONT_REGULAR, 8)
    y_top = PAGE_H - 35 * mm
    c.drawString(15 * mm, y_top, "Сумма страховых взносов, выплаченных работникам пособий по временной нетрудоспособности")
    c.drawString(15 * mm, y_top - 3.5 * mm, "и платежей (взносов) по договорам добровольного личного страхования (нарастающим итогом),")
    c.drawString(15 * mm, y_top - 7 * mm, "предусмотренных пунктом 3.1 статьи 346.21 НК РФ, уменьшающая сумму исчисленного")
    c.drawString(15 * mm, y_top - 10.5 * mm, "за налоговый (отчетный) период налога (авансового платежа по налогу):")

    y_header = PAGE_H - 56 * mm
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, y_header, "Показатели")
    c.drawString(95 * mm, y_header, "Код строки")
    c.drawString(130 * mm, y_header, "Значения показателей (в рублях)")

    rows = [("Сумма страховых взносов, выплаченных работникам пособий\nпо временной нетрудоспособности и платежей (взносов) по договорам\nдобровольного личного страхования за первый квартал", "140", s211.get("line_140"), "money"), ("Сумма страховых взносов за полугодие", "141", s211.get("line_141"), "money"), ("Сумма страховых взносов за девять месяцев", "142", s211.get("line_142"), "money"), ("Сумма страховых взносов за налоговый период", "143", s211.get("line_143"), "money")]

    y = y_header - 10 * mm
    for label, code, value, typ in rows:
        c.setFont(_FONT_REGULAR, 7)
        raw_lines: List[str] = []
        for part in label.split("\n"):
            raw_lines.extend(_wrap_line(part, 52))
        for j, line in enumerate(raw_lines):
            c.drawString(15 * mm, y - j * 3.5 * mm, line)
        _draw_cells(c, 95 * mm, y - 2 * mm, 3, code, cell_w=4.5 * mm, cell_h=5.5 * mm, font_size=9)
        _draw_money_cells(c, 130 * mm, y - 2 * mm, value)
        y -= max(len(raw_lines) * 3.5 + 4, 10) * mm

    c.setFont(_FONT_REGULAR, 8)
    c.drawString(15 * mm, 30 * mm, "Достоверность и полноту сведений, указанных на данной странице, подтверждаю:")
    c.drawString(15 * mm, 20 * mm, "Подпись _______________________")
    c.drawString(90 * mm, 20 * mm, "Дата [__.__.__]")

def generate_pdf(
    decl_data: Dict[str, Any],
    project_data: Dict[str, Any],
) -> bytes:
    """
    Сформировать 4-страничный PDF-декларацию КНД 1152017.

    С апреля 2026 г. использует подложку из официального пустого бланка ФНС
    (pdf_overlay.generate_pdf), чтобы визуально совпадать с формой 1-в-1.
    Старая реализация «с нуля» (_draw_title_page и т.д.) оставлена в файле
    как резервный путь и для справочных целей.

    Args:
        decl_data: результат TaxEngine.get_declaration_data()
        project_data: {inn, kpp?, fio, tax_period_year, oktmo, ifns_code, phone?}

    Returns:
        bytes PDF (4 страницы: Титульный + Раздел 1.1 + Раздел 2.1.1 ч.1 + ч.2)
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError(
            "reportlab не установлен. PDF через старый генератор недоступен. "
            "Используйте основной мастер (wizard) — он формирует PDF через LibreOffice."
        )

    # Основной путь — overlay поверх официального бланка ФНС
    try:
        from app.services import pdf_overlay
        return pdf_overlay.generate_pdf(decl_data, project_data)
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning(
            "pdf_overlay не сработал (%s) — используем резервный рендер",
            exc,
        )

    # -------- резервный путь (старый рендер) --------
    global _FONT_REGULAR, _FONT_BOLD
    _FONT_REGULAR, _FONT_BOLD = _register_cyrillic_font()

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    c.setTitle("Налоговая декларация УСН (КНД 1152017)")
    c.setAuthor("Tax Declaration Generator")

    inn = str(project_data.get("inn", "") or "").strip()
    kpp = str(project_data.get("kpp", "") or "").strip()  # Для ИП всегда пусто
    fio = str(project_data.get("fio", "") or "").strip()
    year = str(project_data.get("tax_period_year",
                                decl_data.get("settings", {}).get("year", "")) or "")
    ifns = str(project_data.get("ifns_code", "") or "").strip()
    phone = str(project_data.get("phone", "") or "").strip()

    _draw_title_page(c, inn, kpp, fio, year, ifns, phone)
    c.showPage()
    s11 = decl_data.get("section_1_1", {})
    _draw_section_1_1(c, inn, kpp, s11)
    c.showPage()
    s211 = decl_data.get("section_2_1_1", {})
    _draw_section_2_1_1_part1(c, inn, kpp, s211)
    c.showPage()
    _draw_section_2_1_1_part2(c, inn, kpp, s211)
    c.showPage()
    c.save()
    return buf.getvalue()


def generate_xml(
    decl_data: Dict[str, Any],
    project_data: Dict[str, Any],
    submission_date=None,
):
    """
    Сформировать XML-файл декларации по формату ФНС 5.07 (КНД 1152017).

    Возвращает tuple (file_id, xml_bytes). xml_bytes закодирован в windows-1251
    и прошёл валидацию по XSD-схеме ФНС.
    """
    from app.services import xml_generator
    return xml_generator.generate_xml(decl_data, project_data, submission_date)


# =============================================================================
# XLSX GENERATION
# =============================================================================
def _fmt_rub(value) -> str:
    """Форматирует целое число рублей с разделителями тысяч."""
    if value is None:
        return "—"
    try:
        n = int(Decimal(str(value)))
    except Exception:
        return str(value)
    return f"{n:,}".replace(",", " ")


def generate_xlsx(
    decl_data: Dict[str, Any],
    project_data: Dict[str, Any],
) -> bytes:
    """
    Сформировать XLSX-декларацию КНД 1152017 (редактируемую).
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    year = project_data.get("tax_period_year", decl_data.get("settings", {}).get("year", ""))
    inn = project_data.get("inn", "")
    fio = project_data.get("fio", "")
    oktmo = project_data.get("oktmo", "")
    ifns = project_data.get("ifns_code", "")

    s211 = decl_data.get("section_2_1_1", {})
    s11 = decl_data.get("section_1_1", {})
    summary = decl_data.get("summary", {})
    settings = decl_data.get("settings", {})

    # ---------- Лист «Титульный» ----------
    ws0 = wb.create_sheet("Титульный лист", 0)
    ws0.merge_cells("A1:C1")
    cc = ws0["A1"]
    cc.value = "Налоговая декларация по УСН (форма КНД 1152017)"
    cc.font = Font(bold=True, size=14)
    cc.alignment = Alignment(horizontal="center")

    ws0.merge_cells("A2:C2")
    ws0["A2"] = (
        "по налогу, уплачиваемому в связи с применением упрощённой "
        "системы налогообложения"
    )
    ws0["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws0.row_dimensions[2].height = 30

    title_rows = [
        ("ИНН", inn),
        ("Налогоплательщик (ФИО)", fio),
        ("Налоговый период (код)", "34"),
        ("Отчётный год", year),
        ("Представляется в налоговый орган", ifns or "—"),
        ("Код по ОКТМО", oktmo or "—"),
        ("Объект налогообложения", "Доходы (код 1)"),
        ("Налоговая ставка, %", settings.get("tax_rate", "6")),
        ("Признак налогоплательщика",
         "1 — производит выплаты физ. лицам" if settings.get("has_employees")
         else "2 — не производит выплаты физ. лицам"),
    ]
    for i, (k, v) in enumerate(title_rows, start=4):
        ws0.cell(row=i, column=1, value=k).font = bold
        ws0.cell(row=i, column=1).border = border
        ws0.cell(row=i, column=2, value=v).border = border
        ws0.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)

    ws0.column_dimensions["A"].width = 40
    ws0.column_dimensions["B"].width = 35
    ws0.column_dimensions["C"].width = 20

    # ---------- Лист «Раздел 1.1» ----------
    ws1 = wb.create_sheet("Раздел 1.1", 1)
    ws1["A1"] = "Раздел 1.1. Сумма налога (авансового платежа), подлежащая уплате (уменьшению)"
    ws1["A1"].font = Font(bold=True, size=12)
    ws1.merge_cells("A1:C1")

    headers = ["Код строки", "Показатель", "Значение, ₽"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    s11_rows = [
        ("010", "Код по ОКТМО", s11.get("line_010", oktmo)),
        ("020", "Сумма авансового платежа к уплате за 1 квартал", s11.get("line_020")),
        ("030", "Код по ОКТМО (при смене)", s11.get("line_030", "")),
        ("040", "Авансовый платёж к уплате за полугодие", s11.get("line_040")),
        ("050", "Авансовый платёж к уменьшению за полугодие", s11.get("line_050")),
        ("060", "Код по ОКТМО (при смене)", s11.get("line_060", "")),
        ("070", "Авансовый платёж к уплате за 9 месяцев", s11.get("line_070")),
        ("080", "Авансовый платёж к уменьшению за 9 месяцев", s11.get("line_080")),
        ("090", "Код по ОКТМО (при смене)", s11.get("line_090", "")),
        ("100", "Сумма налога к доплате за налоговый период", s11.get("line_100")),
        ("110", "Сумма налога к уменьшению за налоговый период", s11.get("line_110")),
    ]

    for i, (code, name, val) in enumerate(s11_rows, start=4):
        ws1.cell(row=i, column=1, value=code).border = border
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=i, column=2, value=name).border = border
        v_cell = ws1.cell(row=i, column=3)
        try:
            v_cell.value = int(Decimal(str(val))) if val not in (None, "", "—") else val
            v_cell.number_format = "# ### ##0"
        except Exception:
            v_cell.value = val
        v_cell.border = border
        v_cell.alignment = Alignment(horizontal="right")
        if code in ("100", "110"):
            for c_idx in range(1, 4):
                ws1.cell(row=i, column=c_idx).fill = total_fill
                ws1.cell(row=i, column=c_idx).font = bold

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 60
    ws1.column_dimensions["C"].width = 20

    # ---------- Лист «Раздел 2.1.1» ----------
    ws2 = wb.create_sheet("Раздел 2.1.1", 2)
    ws2["A1"] = "Раздел 2.1.1. Расчёт налога (УСН Доходы)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.merge_cells("A1:C1")

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    s211_rows = [
        ("101", "Код признака применения ставки", s211.get("line_101", 1)),
        ("102", "Признак налогоплательщика (1/2)", s211.get("line_102", 2)),
        ("110", "Доходы за 1 квартал (нараст. итогом)", s211.get("line_110")),
        ("111", "Доходы за полугодие", s211.get("line_111")),
        ("112", "Доходы за 9 месяцев", s211.get("line_112")),
        ("113", "Доходы за налоговый период", s211.get("line_113")),
        ("120", "Ставка налога, % (Q1) ×10", s211.get("line_120")),
        ("121", "Ставка налога, % (HY) ×10", s211.get("line_121")),
        ("122", "Ставка налога, % (9m) ×10", s211.get("line_122")),
        ("123", "Ставка налога, % (год) ×10", s211.get("line_123")),
        ("130", "Исчисленный налог за 1 квартал", s211.get("line_130")),
        ("131", "Исчисленный налог за полугодие", s211.get("line_131")),
        ("132", "Исчисленный налог за 9 месяцев", s211.get("line_132")),
        ("133", "Исчисленный налог за налоговый период", s211.get("line_133")),
        ("140", "Взносы, уменьшающие налог за 1 квартал", s211.get("line_140")),
        ("141", "Взносы, уменьшающие налог за полугодие", s211.get("line_141")),
        ("142", "Взносы, уменьшающие налог за 9 месяцев", s211.get("line_142")),
        ("143", "Взносы, уменьшающие налог за налоговый период", s211.get("line_143")),
    ]
    for i, (code, name, val) in enumerate(s211_rows, start=4):
        ws2.cell(row=i, column=1, value=code).border = border
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=2, value=name).border = border
        v_cell = ws2.cell(row=i, column=3)
        try:
            v_cell.value = int(Decimal(str(val))) if val not in (None, "", "—") else val
            v_cell.number_format = "# ### ##0"
        except Exception:
            v_cell.value = val
        v_cell.border = border
        v_cell.alignment = Alignment(horizontal="right")
        if code in ("113", "133", "143"):
            for c_idx in range(1, 4):
                ws2.cell(row=i, column=c_idx).fill = total_fill
                ws2.cell(row=i, column=c_idx).font = bold

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 20

    # ---------- Лист «Справка» ----------
    ws3 = wb.create_sheet("Справка", 3)
    ws3["A1"] = "Справочная информация (не входит в форму декларации)"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3.merge_cells("A1:B1")

    ref_rows = [
        ("Общая сумма дохода за год", summary.get("total_income")),
        ("Налог исчисленный за год", summary.get("total_tax_calculated")),
        ("Взносы применённые за год", summary.get("total_contributions_applied")),
        ("Налог после уменьшения за год", summary.get("total_tax_after_reduction")),
        ("Сумма расчётных авансов (Q1+HY+9m)", summary.get("total_advances_due_calc")),
        ("К уплате за год (стр. 100)", summary.get("final_tax_due")),
        ("К уменьшению за год (стр. 110)", summary.get("overpayment")),
        ("1% с превышения 300 000 ₽ (расчётный)", decl_data.get("one_percent_calculated", 0)),
    ]
    for i, (name, val) in enumerate(ref_rows, start=3):
        ws3.cell(row=i, column=1, value=name).border = border
        ws3.cell(row=i, column=1).font = bold
        v_cell = ws3.cell(row=i, column=2)
        try:
            v_cell.value = int(Decimal(str(val))) if val not in (None, "", "—") else val
            v_cell.number_format = "# ### ##0"
        except Exception:
            v_cell.value = val
        v_cell.border = border
        v_cell.alignment = Alignment(horizontal="right")

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
