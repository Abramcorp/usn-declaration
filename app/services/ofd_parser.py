"""
OFD (Оператор Фискальных Данных) receipt file parser.

Supports:
  - xlsx (openpyxl)
  - xls  (xlrd, если установлен; иначе — ошибка с подсказкой)
  - csv  (csv модуль, авто-определение разделителя)

Standard ОФД export columns:
- РН, Место расчетов, Касса, Дата ФД, Тип ФД, Номер ФД,
  Признак расчета, Сумма чека, Наличные, Безналичные, Ошибки ФЛК

The parser:
1. Reads only "Кассовый чек" rows (skipping shift open/close reports).
2. Splits every receipt into its cash-part and card-part (a mixed receipt
   becomes two logical rows, one per payment type).
3. Marks "Приход" as sale and "Возврат прихода" as refund.

It also provides helpers to aggregate receipts by day/payment-type, which
is consumed by the revenue-reconciliation logic (OFD vs. bank acquiring).
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import csv
import io


# Columns we expect (tolerant mapping — case/spacing insensitive)
# Расширенный список алиасов для разных ОФД-операторов
COL_ALIASES: Dict[str, Tuple[str, ...]] = {
    "rn":            ("рн", "rn", "регистрационный номер"),
    "location":      ("место расчетов", "место расчётов", "адрес", "торговая точка",
                      "место расчёта", "место расчета"),
    "kkt":           ("касса", "номер ккт", "ккт", "заводской номер ккт",
                      "зав. номер ккт", "серийный номер"),
    "fd_date":       ("дата фд", "дата", "дата и время", "дата документа",
                      "дата чека", "дата/время", "дата (мск)", "дата операции"),
    "fd_type":       ("тип фд", "тип документа", "вид документа", "тип"),
    "fd_number":     ("номер фд", "номер", "фд", "№ фд", "номер документа",
                      "номер чека", "№"),
    "operation":     ("признак расчета", "признак расчёта", "тип операции",
                      "операция", "тип расчёта", "тип расчета"),
    "total":         ("сумма чека", "итог", "итого", "сумма", "всего",
                      "сумма расчёта", "сумма расчета", "общая сумма"),
    "cash":          ("наличные", "наличными", "нал", "сумма наличными"),
    "card":          ("безналичные", "электронные", "безнал", "банковская карта",
                      "сумма безналичными", "эл. средства платежа"),
    "errors":        ("ошибки флк", "ошибки"),
}

# Максимальное число строк для поиска заголовка
MAX_HEADER_SEARCH_ROWS = 20


class OfdParser:
    """Parse ОФД receipt exports (xlsx, xls, csv)."""

    RECEIPT_TYPE = "Кассовый чек"

    # Варианты значений "Тип ФД" которые считаем кассовыми чеками
    RECEIPT_TYPE_ALIASES = (
        "кассовый чек",
        "чек",
        "кассовый чек (приход)",
        "кассовый чек (возврат прихода)",
    )

    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse an ОФД file (xlsx, xls, or csv).

        Returns:
            {
                "receipts": [
                    {
                        "receipt_date": datetime,
                        "amount": Decimal,
                        "payment_type": "cash"|"card",
                        "operation_type": "sale"|"refund",
                        "receipt_number": str,
                        "kkt_number": str,
                        "point_of_sale": str,
                    }, ...
                ],
                "total_receipts": int,
                "total_cash": Decimal,
                "total_card": Decimal,
                "total_refund_cash": Decimal,
                "total_refund_card": Decimal,
                "period_start": date|None,
                "period_end": date|None,
                "errors": [str],
                "warnings": [str],
            }
        """
        result: Dict[str, Any] = {
            "receipts": [],
            "total_receipts": 0,
            "total_cash": Decimal("0"),
            "total_card": Decimal("0"),
            "total_refund_cash": Decimal("0"),
            "total_refund_card": Decimal("0"),
            "period_start": None,
            "period_end": None,
            "errors": [],
            "warnings": [],
        }

        path = Path(file_path)
        if not path.exists():
            result["errors"].append(f"Файл не найден: {file_path}")
            return result

        suffix = path.suffix.lower()

        if suffix == '.csv':
            rows, error = self._read_csv(path)
            if error:
                result["errors"].append(error)
                return result
        elif suffix == '.xls':
            rows, error = self._read_xls(path)
            if error:
                result["errors"].append(error)
                return result
        elif suffix == '.xlsx':
            rows, error = self._read_xlsx(path)
            if error:
                result["errors"].append(error)
                return result
        else:
            result["errors"].append(
                f"Неподдерживаемый формат файла: {suffix}. "
                "Поддерживаются: .xlsx, .xls, .csv"
            )
            return result

        if not rows:
            result["errors"].append("Файл пуст или не содержит данных.")
            return result

        # Найти строку заголовков
        header_idx, col_map = self._find_header_and_map(rows)
        if header_idx is None:
            result["errors"].append(
                "Не найдены заголовки таблицы. Ожидаются колонки: "
                "'Дата ФД', 'Тип ФД', 'Признак расчета', 'Сумма чека' и т.д. "
                "Проверьте формат экспорта ОФД."
            )
            return result

        has_fd_type = "fd_type" in col_map
        if "operation" not in col_map and "total" not in col_map:
            result["errors"].append(
                "В файле нет обязательных колонок. Ожидаются: "
                "'Признак расчета' и 'Сумма чека'. "
                f"Найденные колонки: {list(col_map.keys())}"
            )
            return result

        min_date: Optional[date] = None
        max_date: Optional[date] = None

        data_rows = rows[header_idx + 1:]
        skipped = 0

        for row in data_rows:
            # Пропускаем пустые строки
            if not any(c for c in row if c is not None and str(c).strip()):
                continue

            # Фильтрация по типу ФД (если колонка есть)
            if has_fd_type:
                fd_type = self._cell(row, col_map, "fd_type")
                if not fd_type:
                    continue
                fd_type_lower = str(fd_type).strip().lower()
                if not any(alias in fd_type_lower or fd_type_lower in alias
                           for alias in self.RECEIPT_TYPE_ALIASES):
                    skipped += 1
                    continue

            operation = self._cell(row, col_map, "operation")
            op_type = self._classify_operation(operation)
            if op_type is None:
                if operation:
                    result["warnings"].append(f"Пропущен чек с признаком '{operation}'")
                continue

            raw_date = self._cell(row, col_map, "fd_date")
            receipt_dt = self._parse_datetime(raw_date)
            if receipt_dt is None:
                result["warnings"].append(f"Не удалось распарсить дату: {raw_date}")
                continue

            d = receipt_dt.date()
            if min_date is None or d < min_date:
                min_date = d
            if max_date is None or d > max_date:
                max_date = d

            cash_amount = self._parse_decimal(self._cell(row, col_map, "cash")) or Decimal("0")
            card_amount = self._parse_decimal(self._cell(row, col_map, "card")) or Decimal("0")
            total_amount = self._parse_decimal(self._cell(row, col_map, "total")) or Decimal("0")

            # Если колонок нал/безнал нет или обе пустые, но есть сумма чека —
            # проверяем наличие колонок: если колонки cash/card не найдены в файле,
            # записываем всю сумму чека по total (считаем что это общий приход).
            # Если колонки есть, но пустые — фолбэк на total как card (эквайринг).
            if cash_amount == 0 and card_amount == 0 and total_amount > 0:
                if "cash" not in col_map and "card" not in col_map:
                    # Нет разбивки вообще — считаем наличными (для налоговой базы)
                    cash_amount = total_amount
                else:
                    # Колонки есть, но пустые — вероятно безнал
                    card_amount = total_amount

            fd_number = str(self._cell(row, col_map, "fd_number") or "")
            kkt = str(self._cell(row, col_map, "kkt") or "")
            location = str(self._cell(row, col_map, "location") or "")

            # Emit one logical row per payment type present
            if cash_amount > 0:
                result["receipts"].append({
                    "receipt_date": receipt_dt,
                    "amount": cash_amount,
                    "payment_type": "cash",
                    "operation_type": op_type,
                    "receipt_number": fd_number,
                    "kkt_number": kkt,
                    "point_of_sale": location,
                })
                if op_type == "sale":
                    result["total_cash"] += cash_amount
                else:
                    result["total_refund_cash"] += cash_amount

            if card_amount > 0:
                result["receipts"].append({
                    "receipt_date": receipt_dt,
                    "amount": card_amount,
                    "payment_type": "card",
                    "operation_type": op_type,
                    "receipt_number": fd_number,
                    "kkt_number": kkt,
                    "point_of_sale": location,
                })
                if op_type == "sale":
                    result["total_card"] += card_amount
                else:
                    result["total_refund_card"] += card_amount

        result["total_receipts"] = len(result["receipts"])
        result["period_start"] = min_date
        result["period_end"] = max_date

        if skipped > 0:
            result["warnings"].append(
                f"Пропущено {skipped} записей (не являются кассовыми чеками)"
            )

        if not result["receipts"] and not result["errors"]:
            result["warnings"].append(
                "Не найдено ни одного кассового чека. "
                "Проверьте, что в файле есть строки с 'Тип ФД' = 'Кассовый чек'."
            )

        return result

    # -------------------- file readers --------------------

    def _read_xlsx(self, path: Path) -> Tuple[Optional[List[list]], Optional[str]]:
        """Read xlsx file into list of rows."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None, "Библиотека openpyxl не установлена."

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            return None, f"Не удалось открыть xlsx-файл: {exc}"

        # Ищем лист с данными (первый, где есть заголовки)
        best_rows = None
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            if rows:
                # Проверяем, есть ли заголовки
                header_idx, col_map = self._find_header_and_map(rows)
                if header_idx is not None:
                    wb.close()
                    return rows, None
                if best_rows is None:
                    best_rows = rows

        wb.close()
        return best_rows, None if best_rows else "Файл xlsx пуст."

    def _read_xls(self, path: Path) -> Tuple[Optional[List[list]], Optional[str]]:
        """Read old .xls format via xlrd."""
        try:
            import xlrd
        except ImportError:
            return None, (
                "Формат .xls не поддерживается без библиотеки xlrd. "
                "Пожалуйста, сохраните файл в формате .xlsx или .csv."
            )

        try:
            wb = xlrd.open_workbook(str(path))
        except Exception as exc:
            return None, f"Не удалось открыть xls-файл: {exc}"

        best_rows = None
        for sheet in wb.sheets():
            rows = []
            for rx in range(sheet.nrows):
                row_values = []
                for cx in range(sheet.ncols):
                    cell = sheet.cell(rx, cx)
                    # Convert xlrd date cells
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                            row_values.append(datetime(*dt_tuple))
                        except Exception:
                            row_values.append(cell.value)
                    else:
                        row_values.append(cell.value)
                rows.append(row_values)
            if rows:
                header_idx, col_map = self._find_header_and_map(rows)
                if header_idx is not None:
                    return rows, None
                if best_rows is None:
                    best_rows = rows

        return best_rows, None if best_rows else "Файл xls пуст."

    def _read_csv(self, path: Path) -> Tuple[Optional[List[list]], Optional[str]]:
        """Read CSV file with auto-detected delimiter."""
        try:
            # Try different encodings
            raw = None
            for enc in ('utf-8-sig', 'utf-8', 'cp1251', 'windows-1251', 'latin-1'):
                try:
                    raw = path.read_text(encoding=enc)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue

            if raw is None:
                return None, "Не удалось определить кодировку CSV-файла."

            # Auto-detect delimiter
            sniffer = csv.Sniffer()
            dialect = None
            delimiter = ';'  # default for Russian CSV
            try:
                dialect = sniffer.sniff(raw[:4096])
                delimiter = dialect.delimiter
            except csv.Error:
                # Fallback: try common delimiters
                delimiter = ';' if raw.count(';') > raw.count(',') else ','

            reader = csv.reader(io.StringIO(raw), delimiter=delimiter)

            rows = [row for row in reader]
            if not rows:
                return None, "CSV-файл пуст."
            return rows, None

        except Exception as exc:
            return None, f"Ошибка чтения CSV: {exc}"

    # -------------------- header detection --------------------

    def _find_header_and_map(self, rows: List[list]) -> Tuple[Optional[int], Dict[str, int]]:
        """Find header row and build column map from raw rows."""
        wanted = {alias for names in COL_ALIASES.values() for alias in names}

        search_limit = min(MAX_HEADER_SEARCH_ROWS, len(rows))
        best_idx = None
        best_map: Dict[str, int] = {}
        best_score = 0

        for row_idx in range(search_limit):
            row = rows[row_idx]
            normalized = [str(c).strip().lower() if c is not None else "" for c in row]

            # Сколько алиасов нашли в этой строке
            matches = sum(1 for cell in normalized if cell in wanted)
            if matches > best_score:
                # Build col_map
                col_map: Dict[str, int] = {}
                # Точное совпадение
                for logical, aliases in COL_ALIASES.items():
                    for i, cell in enumerate(normalized):
                        if cell in aliases:
                            col_map[logical] = i
                            break

                # Fuzzy: если ячейка == алиас как подстрока (но только для не найденных ещё)
                for logical, aliases in COL_ALIASES.items():
                    if logical in col_map:
                        continue
                    for i, cell in enumerate(normalized):
                        if not cell or len(cell) < 3:
                            continue
                        if i in col_map.values():
                            continue  # колонка уже занята
                        for alias in aliases:
                            if cell == alias or alias == cell:
                                col_map[logical] = i
                                break
                        if logical in col_map:
                            break

                if col_map:
                    best_idx = row_idx
                    best_map = col_map
                    best_score = matches

        if best_idx is not None and best_score >= 2:
            return best_idx, best_map
        return None, {}

    # -------------------- helpers --------------------

    @staticmethod
    def _cell(row: list, col_map: Dict[str, int], key: str):
        idx = col_map.get(key)
        if idx is None:
            return None
        if idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _classify_operation(value) -> Optional[str]:
        """Map 'Признак расчета' to sale/refund."""
        if value is None:
            return None
        s = str(value).strip().lower()
        if not s:
            return None
        # Возврат — проверяем первым, т.к. "возврат прихода" содержит "приход"
        if "возврат" in s:
            return "refund"
        if s == "приход" or s.startswith("приход"):
            return "sale"
        # Дополнительные варианты
        if s in ("продажа", "реализация"):
            return "sale"
        return None

    @staticmethod
    def _parse_datetime(value) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        s = str(value).strip()
        if not s:
            return None
        for fmt in (
            "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_decimal(value) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None
        s = str(value).strip().replace("\xa0", "").replace(" ", "")
        if not s:
            return None
        s = s.replace(",", ".")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None


def parse_ofd_file(file_path: str) -> Dict[str, Any]:
    """Universal parser — auto-detects format by extension."""
    return OfdParser().parse(file_path)


# Backward compatibility
def parse_ofd_xlsx(file_path: str) -> Dict[str, Any]:
    """Backward-compatible wrapper (now supports xlsx, xls, csv)."""
    return OfdParser().parse(file_path)


# ---------------------------------------------------------------------------
# Reconciliation: compare OFD card totals with bank acquiring deposits per day
# ---------------------------------------------------------------------------

def aggregate_by_day(receipts: List[Dict[str, Any]]) -> Dict[date, Dict[str, Decimal]]:
    """Aggregate parsed OFD receipts by calendar day.

    Returns: { day: {"cash": Decimal, "card": Decimal,
                      "refund_cash": Decimal, "refund_card": Decimal} }
    """
    out: Dict[date, Dict[str, Decimal]] = defaultdict(lambda: {
        "cash": Decimal("0"),
        "card": Decimal("0"),
        "refund_cash": Decimal("0"),
        "refund_card": Decimal("0"),
    })
    for r in receipts:
        d = r["receipt_date"].date() if isinstance(r["receipt_date"], datetime) else r["receipt_date"]
        key = ("refund_" if r["operation_type"] == "refund" else "") + r["payment_type"]
        out[d][key] += r["amount"]
    return dict(out)


def reconcile_daily(
    ofd_by_day: Dict[date, Dict[str, Decimal]],
    bank_acquiring_by_day: Dict[date, Decimal],
) -> List[Dict[str, Any]]:
    """Merge daily OFD totals with bank acquiring deposits.

    For each day returns:
        {
          "date": date,
          "ofd_cash": Decimal,
          "ofd_card": Decimal,
          "ofd_refund_cash": Decimal,
          "ofd_refund_card": Decimal,
          "bank_acquiring": Decimal,
          "acquiring_diff": ofd_card - bank_acquiring,
          "cash_revenue": ofd_cash - ofd_refund_cash,
        }
    """
    days = sorted(set(ofd_by_day.keys()) | set(bank_acquiring_by_day.keys()))
    out: List[Dict[str, Any]] = []
    for d in days:
        o = ofd_by_day.get(d, {})
        ofd_cash = o.get("cash", Decimal("0"))
        ofd_card = o.get("card", Decimal("0"))
        ref_cash = o.get("refund_cash", Decimal("0"))
        ref_card = o.get("refund_card", Decimal("0"))
        bank = bank_acquiring_by_day.get(d, Decimal("0"))
        out.append({
            "date": d,
            "ofd_cash": ofd_cash,
            "ofd_card": ofd_card,
            "ofd_refund_cash": ref_cash,
            "ofd_refund_card": ref_card,
            "bank_acquiring": bank,
            "acquiring_diff": ofd_card - ref_card - bank,
            "cash_revenue": ofd_cash - ref_cash,
        })
    return out
