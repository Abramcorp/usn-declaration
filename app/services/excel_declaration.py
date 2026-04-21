"""
Fill Russian ИП УСН declaration (КНД 1152017) Excel template.

Supports two form versions:
  - 2024 form (приказ ФНС от 02.10.2024): 11 sheets, quarterly breakdown
  - 2025 form (приказ Минфина от 22.06.2009 №58н): 3 sheets, annual totals
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
import shutil

DATA_DIR = Path(__file__).resolve().parent.parent.parent / 'data'


def _split_fio(fio_str: str) -> list:
    """Разбить ФИО на [Фамилия, Имя, Отчество].

    Отчество может состоять из нескольких слов (например «ШАКИР ОГЛЫ»),
    поэтому берём: слово 1 = фамилия, слово 2 = имя,
    всё остальное = отчество (одной строкой).
    """
    words = fio_str.strip().split()
    if len(words) <= 3:
        return words  # 1, 2 или 3 слова — всё просто
    # 4+ слов: фамилия, имя, остальное — отчество
    return [words[0], words[1], ' '.join(words[2:])]

TEMPLATES = {
    2024: DATA_DIR / 'declaration_template_2024.xlsx',
    2025: DATA_DIR / 'declaration_template_2025.xlsx',
}
# Fallback: old template for years without a specific template
TEMPLATE_XLSX = DATA_DIR / 'declaration_template.xlsx'


# =====================================================================
#  Helpers
# =====================================================================

_FONT = Font(name='Arial', size=11)


def _write_char(ws, coord, ch, font=_FONT):
    cell = ws[coord]
    cell.value = ch
    cell.font = font
    cell.alignment = Alignment(horizontal='center', vertical='center')


def write_chars(ws, row, cols, text, font=_FONT, align='left', pad_char='', keep_spaces=False):
    """Write each character of *text* into cells (row, cols[i])."""
    s = str(text) if text is not None else ''
    n = len(cols)
    if align == 'right':
        chars = list(s.rjust(n, pad_char) if pad_char else s.rjust(n))[-n:]
    else:
        chars = list(s)
        while len(chars) < n:
            chars.append(pad_char if pad_char else None)
    for i in range(n):
        ch = chars[i] if i < len(chars) else None
        if ch is None:
            continue
        if ch == ' ' and not keep_spaces:
            continue
        _write_char(ws, f'{cols[i]}{row}', ch, font)


def _fix_barcode_sizes(wb):
    """Ensure barcode images have reasonable width/height attributes.

    TwoCellAnchor controls the display area, but some viewers (e.g. LibreOffice,
    PDF converters) may use the Image.width/height as the *rendered* size.
    If these are tiny (e.g. 144x98 from low-res templates), barcodes appear
    as small unreadable squares.

    We set minimum 800x450 so that even if a viewer ignores TwoCellAnchor,
    the barcode is still readable.
    """
    MIN_W, MIN_H = 800, 450
    for ws in wb.worksheets:
        for img in ws._images:
            if img.width < MIN_W or img.height < MIN_H:
                img.width = MIN_W
                img.height = MIN_H


def _remove_sheets(wb, keep):
    """Delete sheets not in *keep* set."""
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]


def _reorder_sheets(wb, order):
    """Reorder sheets to match *order* list."""
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != idx:
                wb.move_sheet(name, offset=idx - cur)


def _parse_date(date_val) -> tuple:
    """Return (dd, mm, yyyy) strings from date_val."""
    from datetime import date
    if isinstance(date_val, date):
        return date_val.strftime('%d'), date_val.strftime('%m'), date_val.strftime('%Y')
    s = str(date_val)
    dd, mm, yyyy = s.split('.')
    return dd, mm, yyyy


# =====================================================================
#  2024 form  (приказ ФНС от 02.10.2024)
#  Sheets: Титул, Раздел 1.1, Раздел 2.1.1, Раздел 2.1.1 (продолжение)
# =====================================================================

_2024_TITUL = {
    'inn_row': 1,  'inn_cols': ['Y','AA','AC','AE','AG','AI','AK','AM','AO','AQ','AS','AU'],
    'str_row': 4,  'str_cols': ['AU','AW','AY'],
    'kor_row': 11, 'kor_cols': ['S','U','W'],
    'period_row': 11, 'period_cols': ['BA','BC'],
    'year_row': 11, 'year_cols': ['BU','BW','BY','CA'],
    'ifns_row': 13, 'ifns_cols': ['AA','AC','AE','AG'],
    'pomestu_row': 13, 'pomestu_cols': ['BW','BY','CA'],
    'fio_rows': [15, 17, 19, 21],
    'fio_cols': None,  # auto-detect 40 merged-pair slots A..CB
    'object_row': 29, 'object_col': 'Q',
    'pages_row': 40, 'pages_cols': ['E','G','I'],
    'signer_row': 44, 'signer_col': 'J',
    'date_row': 53,
    'date_day': ['V','X'], 'date_month': ['AB','AD'], 'date_year': ['AH','AJ','AL','AN'],
}

_2024_R11 = {
    'oktmo_010': (13, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']),
    'line_020':  (15, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'oktmo_030': (18, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']),
    'line_040':  (20, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'line_050':  (23, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'oktmo_060': (26, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']),
    'line_070':  (28, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'line_080':  (31, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'oktmo_090': (34, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']),
    'line_100':  (36, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'line_101':  (39, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
    'line_110':  (41, ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']),
}

_2024_R211 = {
    # Признак ставки: 1=6%, 2=8% — single cell
    'line_101': (11, ['Z']),
    # Признак налогоплательщика: 1=с работниками, 2=без
    'line_102': (15, ['Z']),
    # Income
    'line_110': (19, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_111': (21, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_112': (23, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_113': (25, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    # Rate: int at Z, dot in template at AC, decimal at AD
    'line_120_int': (27, ['Z']),  'line_120_dec': (27, ['AD']),
    'line_121_int': (29, ['Z']),  'line_121_dec': (29, ['AD']),
    'line_122_int': (31, ['Z']),  'line_122_dec': (31, ['AD']),
    'line_123_int': (33, ['Z']),  'line_123_dec': (33, ['AD']),
    # Tax
    'line_130': (38, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_131': (40, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_132': (42, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_133': (44, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
}

_2024_R211_CONT = {
    'line_140': (11, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_141': (14, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_142': (17, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_143': (20, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
}


def _fill_2024(template_path, out_path, project_data, decl_data):
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)
    t = _2024_TITUL

    # ---- Титул ----
    ws = wb['Титул']
    write_chars(ws, t['inn_row'], t['inn_cols'], inn)
    write_chars(ws, t['str_row'], t['str_cols'], '001')
    write_chars(ws, t['kor_row'], t['kor_cols'], '0', pad_char='-')
    write_chars(ws, t['period_row'], t['period_cols'],
                decl_data.get('period_code', '34'))
    write_chars(ws, t['year_row'], t['year_cols'],
                str(project_data.get('tax_period_year', 2024)))
    write_chars(ws, t['ifns_row'], t['ifns_cols'],
                str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, t['pomestu_row'], t['pomestu_cols'], '120')

    # ФИО — split by words, 1 word per row
    fio_parts = _split_fio(project_data.get('fio', ''))
    # Build merged-pair FIO cols from first FIO row (auto-detect)
    fio_cols = _get_fio_cols_2024(ws, t['fio_rows'][0])
    for i, row in enumerate(t['fio_rows']):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val, keep_spaces=True)

    _write_char(ws, f"{t['object_col']}{t['object_row']}", '1')
    write_chars(ws, t['pages_row'], t['pages_cols'], '004')
    _write_char(ws, f"{t['signer_col']}{t['signer_row']}", '1')

    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.04.2025'))
    write_chars(ws, t['date_row'], t['date_day'], dd)
    write_chars(ws, t['date_row'], t['date_month'], mm)
    write_chars(ws, t['date_row'], t['date_year'], yyyy)

    # ---- Раздел 1.1 ----
    ws = wb['Раздел 1.1']
    sec11 = decl_data.get('section_1_1', {})
    oktmo = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    for key in ('oktmo_010', 'oktmo_030', 'oktmo_060', 'oktmo_090'):
        row, cols = _2024_R11[key]
        write_chars(ws, row, cols, oktmo)
    for key in ('line_020','line_040','line_050','line_070','line_080',
                'line_100','line_101','line_110'):
        row, cols = _2024_R11[key]
        val = sec11.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Раздел 2.1.1 ----
    ws = wb['Раздел 2.1.1']
    sec211 = decl_data.get('section_2_1_1', {})

    # Признак ставки (1=6%)
    row, cols = _2024_R211['line_101']
    _write_char(ws, f'{cols[0]}{row}', '1')
    # Признак налогоплательщика (1=с работниками, 2=ИП без)
    row, cols = _2024_R211['line_102']
    v102 = str(sec211.get('line_101', '2'))  # old key was line_101
    _write_char(ws, f'{cols[0]}{row}', v102)

    # Income 110-113
    for key in ('line_110','line_111','line_112','line_113'):
        row, cols = _2024_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # Rates 120-123
    for key in ('line_120','line_121','line_122','line_123'):
        val = sec211.get(key, 6.0)
        int_part = str(int(val))
        dec_part = str(int(round((float(val) - int(val)) * 10)))
        ri, ci = _2024_R211[f'{key}_int']
        rd, cd = _2024_R211[f'{key}_dec']
        write_chars(ws, ri, ci, int_part)
        write_chars(ws, rd, cd, dec_part)

    # Tax 130-133
    for key in ('line_130','line_131','line_132','line_133'):
        row, cols = _2024_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Раздел 2.1.1 (продолжение) ----
    ws = wb['Раздел 2.1.1 (продолжение)']
    for key in ('line_140','line_141','line_142','line_143'):
        row, cols = _2024_R211_CONT[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # Keep only relevant sheets
    keep = {'Титул', 'Раздел 1.1', 'Раздел 2.1.1', 'Раздел 2.1.1 (продолжение)'}
    _remove_sheets(wb, keep)
    _reorder_sheets(wb, ['Титул', 'Раздел 1.1', 'Раздел 2.1.1', 'Раздел 2.1.1 (продолжение)'])

    _fix_barcode_sizes(wb)
    wb.save(out_path)
    print(f'Saved 2024 declaration: {out_path} (листы: {wb.sheetnames})')


def _get_fio_cols_2024(ws, row):
    """Auto-detect FIO merged-pair columns for the 2024 form."""
    from openpyxl.utils import get_column_letter
    merges = [mr for mr in ws.merged_cells.ranges if mr.min_row <= row <= mr.max_row]
    bordered = set()
    for c in range(1, min(82, ws.max_column + 1)):
        b = ws.cell(row, c).border
        if any([b.left.style, b.right.style, b.top.style, b.bottom.style]):
            bordered.add(c)
    slots = []
    visited = set()
    for c in sorted(bordered):
        if c in visited:
            continue
        in_merge = False
        for mr in merges:
            if mr.min_col <= c <= mr.max_col:
                slots.append(get_column_letter(mr.min_col))
                for cc in range(mr.min_col, mr.max_col + 1):
                    visited.add(cc)
                in_merge = True
                break
        if not in_merge:
            slots.append(get_column_letter(c))
            visited.add(c)
    return slots


# =====================================================================
#  2025 form  (приказ Минфина от 22.06.2009 №58н)
#  Sheets: стр.1, стр.2_Разд.1, стр.3_Разд.2
# =====================================================================

_2025_TITUL = {
    'inn_row': 3, 'inn_cols': ['AK','AN','AQ','AT','AW','AZ','BC','BF','BI','BL','BO','BR'],
    'str_row': 6, 'str_cols': ['BR','BU','BX'],
    'kor_row': 14, 'kor_cols': ['X','AA','AD'],
    'period_row': 14, 'period_cols': ['BU','BX'],
    'year_row': 14, 'year_cols': ['DD','DG','DJ','DM'],
    'ifns_row': 16, 'ifns_cols': ['AX','BA','BD','BG'],
    'pomestu_row': 16, 'pomestu_cols': ['DF','DI','DL'],
    'fio_rows': [18, 20, 22, 24],
    'fio_cols': None,  # auto-detect
    'pages_row': 31, 'pages_cols': ['E','H','K'],
    'signer_row': 37, 'signer_col': 'M',
    'date_row': 69,
    'date_day': ['AE','AH'], 'date_month': ['AN','AQ'], 'date_year': ['AW','AZ','BC','BF'],
}

_2025_R1 = {
    'line_001': (14, ['BI']),  # объект: 1=доходы
    'okato_010': (16, ['BI','BL','BO','BR','BU','BX','CA','CD','CG','CJ','CM']),
    'kbk_020': (18, ['BI','BL','BO','BR','BU','BX','CA','CD','CG','CJ','CM',
                      'CP','CS','CV','CY','DB','DE','DH','DK','DN']),
    'line_030': (23, ['BI','BL','BO','BR','BU','BX','CA','CD']),
    'line_040': (25, ['BI','BL','BO','BR','BU','BX','CA','CD']),
    'line_050': (27, ['BI','BL','BO','BR','BU','BX','CA','CD']),
    'line_060': (30, ['BI','BL','BO','BR','BU','BX','CA','CD']),
    'line_070': (33, ['BI','BL','BO','BR','BU','BX','CA','CD']),
}

_2025_R2 = {
    # Rate: int at CP,CS; dot at CV in template; dec at CY
    'line_201_int': (14, ['CP','CS']),
    'line_201_dec': (14, ['CY']),
    'line_210': (16, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_220': (18, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_230': (21, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_240': (24, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_250': (27, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_260': (30, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_270': (33, ['CP','CS','CV','CY','DB','DE','DH','DK']),
    'line_280': (37, ['CP','CS','CV','CY','DB','DE','DH','DK']),
}

# КБК УСН «Доходы» (6%) — основной платёж
KBK_USN_INCOME = '18210501011011000110'


def _fill_2025(template_path, out_path, project_data, decl_data):
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)
    t = _2025_TITUL

    # ---- стр.1 (Титул) ----
    ws = wb['стр.1']
    write_chars(ws, t['inn_row'], t['inn_cols'], inn)
    write_chars(ws, t['str_row'], t['str_cols'], '001')
    write_chars(ws, t['kor_row'], t['kor_cols'], '0', pad_char='-')
    write_chars(ws, t['period_row'], t['period_cols'],
                decl_data.get('period_code', '34'))
    write_chars(ws, t['year_row'], t['year_cols'],
                str(project_data.get('tax_period_year', 2025)))
    write_chars(ws, t['ifns_row'], t['ifns_cols'],
                str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, t['pomestu_row'], t['pomestu_cols'], '120')

    # ФИО
    fio_parts = _split_fio(project_data.get('fio', ''))
    fio_cols = _get_fio_cols_2025(ws, t['fio_rows'][0])
    for i, row in enumerate(t['fio_rows']):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val, keep_spaces=True)

    write_chars(ws, t['pages_row'], t['pages_cols'], '003')
    _write_char(ws, f"{t['signer_col']}{t['signer_row']}", '1')

    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.04.2026'))
    write_chars(ws, t['date_row'], t['date_day'], dd)
    write_chars(ws, t['date_row'], t['date_month'], mm)
    write_chars(ws, t['date_row'], t['date_year'], yyyy)

    # ---- стр.2_Разд.1 ----
    ws = wb['стр.2_Разд.1']
    sec1 = decl_data.get('section_1', {})

    # Объект: 1 = доходы
    row, cols = _2025_R1['line_001']
    _write_char(ws, f'{cols[0]}{row}', '1')

    # ОКАТО (используем ОКТМО, дополняя до 11 символов)
    okato = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    row, cols = _2025_R1['okato_010']
    write_chars(ws, row, cols, okato)

    # КБК
    kbk = sec1.get('kbk', KBK_USN_INCOME)
    row, cols = _2025_R1['kbk_020']
    write_chars(ws, row, cols, str(kbk))

    # Авансовые платежи (030/040/050 = вычисленные авансы за Q1/H1/9M)
    for key in ('line_030', 'line_040', 'line_050', 'line_060', 'line_070'):
        row, cols = _2025_R1[key]
        val = sec1.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- стр.3_Разд.2 ----
    ws = wb['стр.3_Разд.2']
    sec2 = decl_data.get('section_2', {})

    # Ставка (6.0 = " 6" + "0")
    rate = sec2.get('rate', 6.0)
    int_part = str(int(rate)).rjust(2)
    dec_part = str(int(round((float(rate) - int(rate)) * 10)))
    ri, ci = _2025_R2['line_201_int']
    rd, cd = _2025_R2['line_201_dec']
    write_chars(ws, ri, ci, int_part)
    write_chars(ws, rd, cd, dec_part)

    # Годовые показатели
    for key in ('line_210', 'line_240', 'line_260', 'line_280'):
        row, cols = _2025_R2[key]
        val = sec2.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # Для объекта «Доходы» строки 220, 230, 250 = 0 (не заполняем)
    # Минимальный налог (стр. 270) = 0 для объекта «Доходы»

    wb.save(out_path)
    print(f'Saved 2025 declaration: {out_path} (листы: {wb.sheetnames})')


def _get_fio_cols_2025(ws, row):
    """Auto-detect FIO merged-triplet columns for the 2025 form."""
    from openpyxl.utils import get_column_letter
    merges = [mr for mr in ws.merged_cells.ranges if mr.min_row <= row <= mr.max_row]
    bordered = set()
    for c in range(1, min(121, ws.max_column + 1)):
        b = ws.cell(row, c).border
        if any([b.left.style, b.right.style, b.top.style, b.bottom.style]):
            bordered.add(c)
    slots = []
    visited = set()
    for c in sorted(bordered):
        if c in visited:
            continue
        in_merge = False
        for mr in merges:
            if mr.min_col <= c <= mr.max_col:
                slots.append(get_column_letter(mr.min_col))
                for cc in range(mr.min_col, mr.max_col + 1):
                    visited.add(cc)
                in_merge = True
                break
        if not in_merge:
            slots.append(get_column_letter(c))
            visited.add(c)
    return slots


# =====================================================================
#  Legacy form (old template for ≤2023)
#  Kept for backward compatibility
# =====================================================================

# Old template cell maps (Титульный лист / Р.1.1 / Р.2.1.1 / Р.2.1.1 (продол.))
_OLD_TITUL_INN_COLS = ['M','N','O','P','Q','R','S','T','U','V','W','X']
_OLD_STR_COLS = ['X','Y','Z']

_OLD_R11 = {
    'oktmo_010': (12, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_020':  (14, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_030': (17, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_040':  (19, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_050':  (22, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_060': (25, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_070':  (27, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_080':  (30, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_090': (33, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_100':  (35, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_101':  (38, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_110':  (41, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
}

_OLD_R211 = {
    'line_101_pr': (12, ['AC']),
    'line_110': (17, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_111': (19, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_112': (21, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_113': (23, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_120_int': (27, ['AC']),  'line_120_dec': (27, ['AE']),
    'line_121_int': (29, ['AC']),  'line_121_dec': (29, ['AE']),
    'line_122_int': (31, ['AC']),  'line_122_dec': (31, ['AE']),
    'line_123_int': (33, ['AC']),  'line_123_dec': (33, ['AE']),
    'line_130': (40, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_131': (43, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_132': (46, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_133': (49, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
}

_OLD_R212 = {
    'line_140': (10, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_141': (14, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_142': (18, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_143': (22, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    # Строки 150, 160, 161, 162 — суммы фиксированных взносов по ст. 430 НК
    'line_150': (28, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_160': (30, ['AB','AC','AD','AE','AF','AG']),
    'line_161': (32, ['AB','AC','AD','AE','AF','AG']),
    'line_162': (34, ['AB','AC','AD','AE','AF','AG']),
}


def _fill_old(template_path, out_path, project_data, decl_data):
    """Fill the old template (pre-2024)."""
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)

    # ---- Титульный лист ----
    ws = wb['Титульный лист']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '001')
    write_chars(ws, 10, ['H','I','J'], '0', pad_char='-')
    write_chars(ws, 10, ['W','X'], decl_data.get('period_code', '34'))
    write_chars(ws, 10, ['AK','AL','AM','AN'], str(project_data.get('tax_period_year', 2024)))
    write_chars(ws, 12, ['N','O','P','Q'], str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, 12, ['AL','AM','AN'], '120')
    fio_parts = _split_fio(project_data.get('fio', ''))
    fio_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
    for i, row in enumerate([14, 16, 18, 20]):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val, keep_spaces=True)
    _write_char(ws, 'I28', '1')
    write_chars(ws, 39, ['C','D','E'], '004')
    _write_char(ws, 'B44', '1')
    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.01.2025'))
    write_chars(ws, 54, ['AE','AF'], dd)
    write_chars(ws, 54, ['AH','AI'], mm)
    write_chars(ws, 54, ['AK','AL','AM','AN'], yyyy)

    # ---- Р.1.1 ----
    ws = wb['Р.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '002')
    sec11 = decl_data.get('section_1_1', {})
    oktmo = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    for key in ('oktmo_010','oktmo_030','oktmo_060','oktmo_090'):
        row, cols = _OLD_R11[key]
        write_chars(ws, row, cols, oktmo)
    for key in ('line_020','line_040','line_050','line_070','line_080',
                'line_100','line_101','line_110'):
        row, cols = _OLD_R11[key]
        val = sec11.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 ----
    ws = wb['Р.2.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '003')
    sec211 = decl_data.get('section_2_1_1', {})
    v101 = str(sec211.get('line_101', '2'))
    _write_char(ws, 'AC12', v101)
    for key in ('line_110','line_111','line_112','line_113'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')
    for key in ('line_120','line_121','line_122','line_123'):
        val = sec211.get(key, 6.0)
        int_part = str(int(val))
        dec_part = str(int(round((float(val) - int(val)) * 10)))
        ri, ci = _OLD_R211[f'{key}_int']
        rd, cd = _OLD_R211[f'{key}_dec']
        write_chars(ws, ri, ci, int_part)
        write_chars(ws, rd, cd, dec_part)
    for key in ('line_130','line_131','line_132','line_133'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 (продол.) ----
    ws = wb['Р.2.1.1 (продол.)']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '004')
    for key in ('line_140','line_141','line_142','line_143'):
        row, cols = _OLD_R212[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # Строки 150, 160, 161, 162 — страховые взносы по ст. 430 НК РФ
    contributions_data = decl_data.get('contributions_430', {})
    for key in ('line_150', 'line_160', 'line_161', 'line_162'):
        if key in _OLD_R212:
            row, cols = _OLD_R212[key]
            val = contributions_data.get(key, 0)
            if val:
                write_chars(ws, row, cols, str(int(val)), pad_char='-')

    keep = {'Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'}
    _remove_sheets(wb, keep)
    _reorder_sheets(wb, ['Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'])

    # Исправить размеры штрихкодов на всех листах
    _fix_barcode_sizes(wb)

    wb.save(out_path)
    print(f'Saved old declaration: {out_path} (листы: {wb.sheetnames})')


# =====================================================================
#  Public API
# =====================================================================

def get_template_for_year(year: int) -> Path:
    """Return the correct template path for the given tax year."""
    if year in TEMPLATES and TEMPLATES[year].exists():
        return TEMPLATES[year]
    return TEMPLATE_XLSX


def fill_declaration(template_path: Path, out_path: Path,
                     project_data: dict, decl_data: dict):
    """Fill declaration using the correct form version based on template."""
    name = template_path.name

    if 'template_2024' in name:
        _fill_2024(template_path, out_path, project_data, decl_data)
    else:
        # 2025 и все остальные годы используют старую форму
        # (Титульный лист / Р.1.1 / Р.2.1.1 / Р.2.1.1 (продол.))
        _fill_old(template_path, out_path, project_data, decl_data)
