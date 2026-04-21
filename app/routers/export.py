"""
Export endpoints for Russian tax declaration system.
Handles export to Excel, CSV, and JSON formats.
"""

import io
import csv
import json
from decimal import Decimal
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.models import Project, BankOperation, TaxCalculation, InsuranceContribution
from app.services.tax_engine import TaxEngine
from app.services.contribution_calculator import calculate_total_ip_contributions
try:
    from app.services.declaration_generator import generate_pdf, generate_xlsx
except ImportError:
    generate_pdf = None
    generate_xlsx = None
from datetime import date

router = APIRouter(prefix="/api/export", tags=["export"])

# Try to import openpyxl, fallback if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _format_decimal(value) -> str:
    """Format Decimal for display."""
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _create_excel_file(project: Project, db: Session) -> io.BytesIO:
    """Create Excel export file with multiple sheets."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl не установлена. Установите зависимость для экспорта в Excel."
        )

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Реестр доходов (Income Registry)
    ws1 = wb.create_sheet("Реестр доходов", 0)
    income_ops = db.query(BankOperation).filter(
        BankOperation.project_id == project.id,
        BankOperation.direction == "income",
        BankOperation.included_in_tax_base == True,
    ).order_by(BankOperation.operation_date).all()

    headers = ["Дата", "Сумма", "Контрагент", "Назначение платежа", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row, op in enumerate(income_ops, 2):
        ws1.cell(row=row, column=1, value=op.operation_date).border = border
        ws1.cell(row=row, column=2, value=float(op.amount)).border = border
        ws1.cell(row=row, column=3, value=op.counterparty or "").border = border
        ws1.cell(row=row, column=4, value=op.purpose).border = border
        ws1.cell(row=row, column=5, value=op.classification).border = border

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 35
    ws1.column_dimensions['E'].width = 15

    # Sheet 2: Исключения (Excluded Operations)
    ws2 = wb.create_sheet("Исключения", 1)
    excluded_ops = db.query(BankOperation).filter(
        BankOperation.project_id == project.id,
        BankOperation.direction == "income",
        BankOperation.classification == "not_income",
    ).order_by(BankOperation.operation_date).all()

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row, op in enumerate(excluded_ops, 2):
        ws2.cell(row=row, column=1, value=op.operation_date).border = border
        ws2.cell(row=row, column=2, value=float(op.amount)).border = border
        ws2.cell(row=row, column=3, value=op.counterparty or "").border = border
        ws2.cell(row=row, column=4, value=op.purpose).border = border
        ws2.cell(row=row, column=5, value=op.exclusion_reason or "").border = border

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = ws1.column_dimensions[col].width

    # Sheet 3: Спорные (Disputed Operations)
    ws3 = wb.create_sheet("Спорные", 2)
    disputed_ops = db.query(BankOperation).filter(
        BankOperation.project_id == project.id,
        BankOperation.classification == "disputed",
    ).order_by(BankOperation.operation_date).all()

    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row, op in enumerate(disputed_ops, 2):
        ws3.cell(row=row, column=1, value=op.operation_date).border = border
        ws3.cell(row=row, column=2, value=float(op.amount)).border = border
        ws3.cell(row=row, column=3, value=op.counterparty or "").border = border
        ws3.cell(row=row, column=4, value=op.purpose).border = border
        ws3.cell(row=row, column=5, value=f"Уверенность: {op.classification_confidence or 0:.0%}").border = border

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws3.column_dimensions[col].width = ws1.column_dimensions[col].width

    # Sheet 4: Расчёт налога (Tax Calculation)
    ws4 = wb.create_sheet("Расчёт налога", 3)
    calculations = db.query(TaxCalculation).filter(
        TaxCalculation.project_id == project.id
    ).all()

    calc_headers = [
        "Период", "Доход нарастающий", "Налог рассчитанный",
        "Взносы применённые", "Налог после снижения", "Авансовые платежи", "Налог к уплате"
    ]

    for col, header in enumerate(calc_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    period_names = {
        "q1": "I квартал",
        "half_year": "За полугодие",
        "nine_months": "За 9 месяцев",
        "year": "За год",
    }

    for row, calc in enumerate(calculations, 2):
        period_name = period_names.get(calc.period, calc.period)
        ws4.cell(row=row, column=1, value=period_name).border = border
        ws4.cell(row=row, column=2, value=float(calc.income_cumulative)).border = border
        ws4.cell(row=row, column=3, value=float(calc.tax_calculated)).border = border
        ws4.cell(row=row, column=4, value=float(calc.contributions_applied)).border = border
        ws4.cell(row=row, column=5, value=float(calc.tax_after_reduction)).border = border
        ws4.cell(row=row, column=6, value=float(calc.advance_paid)).border = border
        ws4.cell(row=row, column=7, value=float(calc.tax_due)).border = border

    for col in range(1, 8):
        ws4.column_dimensions[chr(64 + col)].width = 18

    # Sheet 5: Декларация (Declaration)
    ws5 = wb.create_sheet("Декларация", 4)

    decl_data = [
        ["Информация о налогоплательщике", ""],
        ["ИНН:", project.inn],
        ["ФИО:", project.fio],
        ["Налоговый период:", project.tax_period_year],
        ["", ""],
        ["Результаты расчётов", ""],
    ]

    for row, data in enumerate(decl_data, 1):
        ws5.cell(row=row, column=1, value=data[0])
        ws5.cell(row=row, column=2, value=data[1])

    # Add calculation results
    row_offset = len(decl_data) + 1
    for calc in calculations:
        period_name = period_names.get(calc.period, calc.period)
        ws5.cell(row=row_offset, column=1, value=f"Налог {period_name}:")
        ws5.cell(row=row_offset, column=2, value=float(calc.tax_due))
        row_offset += 1

    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 20

    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


@router.get("/excel/{project_id}")
def export_excel(project_id: int, db: Session = Depends(get_db)):
    """
    Export full calculation to Excel file.

    Creates workbook with sheets:
    - Реестр доходов (Income registry)
    - Исключения (Excluded operations)
    - Спорные (Disputed operations)
    - Расчёт налога (Tax calculation)
    - Декларация (Declaration)

    Args:
        project_id: ID of the project

    Returns:
        Excel file as attachment.

    Raises:
        HTTPException: If project not found or openpyxl not available.
    """
    # Verify project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Проект с ID {project_id} не найден"
        )

    # Create Excel file
    excel_file = _create_excel_file(project, db)

    # Prepare response
    filename = f"Налоговая_декларация_{project.inn}_{project.tax_period_year}.xlsx"
    return StreamingResponse(
        iter([excel_file.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/csv/{project_id}")
def export_csv(project_id: int, db: Session = Depends(get_db)):
    """
    Export income registry as CSV file.

    Args:
        project_id: ID of the project

    Returns:
        CSV file as attachment.

    Raises:
        HTTPException: If project not found.
    """
    # Verify project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Проект с ID {project_id} не найден"
        )

    # Query income operations
    operations = db.query(BankOperation).filter(
        BankOperation.project_id == project_id,
        BankOperation.direction == "income",
        BankOperation.included_in_tax_base == True,
    ).order_by(BankOperation.operation_date).all()

    # Create CSV
    output = io.StringIO()
    fieldnames = [
        "Дата операции", "Сумма", "Контрагент", "ИНН контрагента",
        "Назначение платежа", "Статус"
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for op in operations:
        writer.writerow({
            "Дата операции": op.operation_date.isoformat(),
            "Сумма": _format_decimal(op.amount),
            "Контрагент": op.counterparty or "",
            "ИНН контрагента": op.counterparty_inn or "",
            "Назначение платежа": op.purpose,
            "Статус": op.classification,
        })

    # Prepare response
    filename = f"Доходы_{project.inn}_{project.tax_period_year}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _build_declaration_payload(project: Project, db: Session) -> tuple:
    """
    Построить данные декларации КНД 1152017 для проекта.
    Возвращает (decl_data, project_data) для передачи в generate_pdf/generate_xlsx.
    """
    from sqlalchemy import func as sa_func

    year = project.tax_period_year

    # Квартальные доходы
    quarterly = {"q1": Decimal("0"), "q2": Decimal("0"), "q3": Decimal("0"), "q4": Decimal("0")}
    quarter_ranges = {
        "q1": (date(year, 1, 1), date(year, 3, 31)),
        "q2": (date(year, 4, 1), date(year, 6, 30)),
        "q3": (date(year, 7, 1), date(year, 9, 30)),
        "q4": (date(year, 10, 1), date(year, 12, 31)),
    }
    for q, (s, e) in quarter_ranges.items():
        v = db.query(sa_func.sum(BankOperation.amount)).filter(
            BankOperation.project_id == project.id,
            BankOperation.direction == "income",
            BankOperation.included_in_tax_base == True,
            BankOperation.operation_date >= s,
            BankOperation.operation_date <= e,
        ).scalar() or Decimal("0")
        quarterly[q] = v

    total_income = sum(quarterly.values(), Decimal("0"))
    auto_contrib = calculate_total_ip_contributions(year, total_income)

    # Взносы — из базы, если сохранены пользователем; иначе автоматические
    contribs_db = db.query(InsuranceContribution).filter(
        InsuranceContribution.project_id == project.id
    ).all()
    contribs_dict = {}
    for c in contribs_db:
        contribs_dict[c.contribution_type] = contribs_dict.get(c.contribution_type, Decimal("0")) + c.amount

    engine_contribs = {
        "fixed_ip": contribs_dict.get("fixed_ip", auto_contrib["fixed"]),
        "one_percent": contribs_dict.get("one_percent", auto_contrib["one_percent"]),
        "employee_insurance": contribs_dict.get("employee_insurance", Decimal("0")),
    }

    settings = {
        "has_employees": project.has_employees,
        "uses_ens": project.uses_ens,
        "employee_start_quarter": project.employee_start_quarter,
        "contribution_input_mode": project.contribution_input_mode or "total",
        "year": year,
        "tax_rate": project.tax_rate or "6",
    }
    engine = TaxEngine(settings)
    calc_result = engine.calculate(quarterly, engine_contribs, {})
    decl_data = engine.get_declaration_data(calc_result, {"oktmo": project.oktmo or ""})

    project_data = {
        "inn": project.inn,
        "fio": project.fio,
        "tax_period_year": year,
        "oktmo": project.oktmo or "",
        "ifns_code": project.ifns_code or "",
    }
    return decl_data, project_data


@router.get("/declaration-pdf/{project_id}")
def export_declaration_pdf(project_id: int, db: Session = Depends(get_db)):
    """Сформировать PDF-декларацию по форме КНД 1152017."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"Проект {project_id} не найден")

    decl_data, project_data = _build_declaration_payload(project, db)
    try:
        pdf_bytes = generate_pdf(decl_data, project_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации PDF: {e}")

    filename = f"Декларация_УСН_{project.inn or 'ИП'}_{project.tax_period_year}.pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/declaration-xlsx/{project_id}")
def export_declaration_xlsx(project_id: int, db: Session = Depends(get_db)):
    """Сформировать XLSX-декларацию по форме КНД 1152017 (редактируемая)."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"Проект {project_id} не найден")

    decl_data, project_data = _build_declaration_payload(project, db)
    try:
        xlsx_bytes = generate_xlsx(decl_data, project_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации XLSX: {e}")

    filename = f"Декларация_УСН_{project.inn or 'ИП'}_{project.tax_period_year}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/json/{project_id}")
def export_json(project_id: int, db: Session = Depends(get_db)):
    """
    Export all project data as JSON.

    Args:
        project_id: ID of the project

    Returns:
        JSON file with all project data.

    Raises:
        HTTPException: If project not found.
    """
    # Verify project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Проект с ID {project_id} не найден"
        )

    # Query all related data
    operations = db.query(BankOperation).filter(
        BankOperation.project_id == project_id
    ).all()

    calculations = db.query(TaxCalculation).filter(
        TaxCalculation.project_id == project_id
    ).all()

    # Build export structure
    export_data = {
        "project": {
            "id": project.id,
            "inn": project.inn,
            "fio": project.fio,
            "tax_period_year": project.tax_period_year,
            "status": project.status,
            "tax_rate": project.tax_rate,
            "has_employees": project.has_employees,
            "uses_ens": project.uses_ens,
            "contribution_input_mode": project.contribution_input_mode,
            "oktmo": project.oktmo,
            "ifns_code": project.ifns_code,
        },
        "operations": [
            {
                "id": op.id,
                "operation_date": op.operation_date.isoformat(),
                "posting_date": op.posting_date.isoformat() if op.posting_date else None,
                "amount": _format_decimal(op.amount),
                "direction": op.direction,
                "purpose": op.purpose,
                "counterparty": op.counterparty,
                "counterparty_inn": op.counterparty_inn,
                "classification": op.classification,
                "classification_confidence": op.classification_confidence,
                "included_in_tax_base": op.included_in_tax_base,
                "is_manually_classified": op.is_manually_classified,
            }
            for op in operations
        ],
        "tax_calculations": [
            {
                "period": calc.period,
                "income_cumulative": _format_decimal(calc.income_cumulative),
                "tax_calculated": _format_decimal(calc.tax_calculated),
                "contributions_applied": _format_decimal(calc.contributions_applied),
                "contribution_limit": _format_decimal(calc.contribution_limit),
                "tax_after_reduction": _format_decimal(calc.tax_after_reduction),
                "advance_paid": _format_decimal(calc.advance_paid),
                "tax_due": _format_decimal(calc.tax_due),
            }
            for calc in calculations
        ],
        "export_date": datetime.utcnow().isoformat(),
    }

    # Prepare response
    filename = f"Экспорт_{project.inn}_{project.tax_period_year}.json"
    return StreamingResponse(
        iter([json.dumps(export_data, ensure_ascii=False, indent=2)]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
